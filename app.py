import calendar
import csv
import os
import re
import sqlite3
import uuid
from datetime import date, datetime, timedelta
from functools import wraps
from io import BytesIO, StringIO
from pathlib import Path

from flask import (
    Flask,
    Response,
    abort,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_wtf.csrf import CSRFError, CSRFProtect, generate_csrf
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename


BASE_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = BASE_DIR / "instance"
DATABASE_PATH = INSTANCE_DIR / "clearpath.db"
UPLOAD_DIR = INSTANCE_DIR / "uploads"
EXTRACTION_UPLOAD_DIR = UPLOAD_DIR / "ec_extractions"
DEFAULT_APPROVAL_TYPES = [
    "EC",
    "CTE",
    "CTO",
    "HWM",
    "BMW",
    "Fire NOC",
    "Forest",
    "AERB",
    "CGWA",
]
FREQUENCY_TYPES = ["General", "One-time", "Monthly", "Quarterly", "Half-yearly", "Yearly"]
LIFECYCLE_STATUSES = ["Pending", "Completed", "On hold", "Not applicable"]
STATUS_FILTERS = ["All", "Pending", "Due in 7 days", "Overdue", "Completed", "On hold", "Not applicable"]
CALENDAR_VIEW_OPTIONS = [1, 2, 3]
SUBMISSION_MODE_OPTIONS = [
    "",
    "Portal",
    "Email",
    "Courier",
    "Hand delivery",
    "Post",
    "Meeting / visit",
    "Other",
]
ALLOWED_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".gif", ".webp"}
ALLOWED_IMPORT_EXTENSIONS = {".csv", ".xlsx"}
IMPORT_HEADER_ALIASES = {
    "s. no": "serial_number",
    "s no": "serial_number",
    "serial no": "serial_number",
    "serial number": "serial_number",
    "condition description": "condition_description",
    "condition": "condition_description",
    "description": "condition_description",
    "approval": "approval_type",
    "noc": "approval_type",
    "noc type": "approval_type",
    "approval type": "approval_type",
    "action to be taken": "action_to_be_taken",
    "action": "action_to_be_taken",
    "due date": "due_date",
    "status": "status",
    "frequency": "frequency",
    "type": "frequency",
    "schedule source": "schedule_source",
    "submitted to": "submitted_to",
    "submission mode": "submission_mode",
    "responsible person": "responsible_person",
    "acknowledgment number": "acknowledgment_number",
    "acknowledgement number": "acknowledgment_number",
    "reference no": "acknowledgment_number",
    "reference number": "acknowledgment_number",
    "remarks": "remarks",
    "notes": "remarks",
}
SCHEMA_WHITELIST = {
    "project_approvals": {"issue_date", "expiry_date", "approval_notes"},
    "documents": {"document_title", "version_notes"},
}
APPROVAL_PORTFOLIO_TYPES = {"EC", "CTE", "CTO"}
REPORT_TABS = {"conditions", "compliance-report", "documents", "activity"}
SOURCE_TYPES = {"manual", "import", "ec_extraction"}


secret_key = os.environ.get("SECRET_KEY")
if not secret_key:
    raise RuntimeError("SECRET_KEY environment variable is required before starting ClearPath.")

app = Flask(__name__, instance_path=str(INSTANCE_DIR), instance_relative_config=True)
app.config["SECRET_KEY"] = secret_key
app.config["DATABASE"] = str(DATABASE_PATH)
app.config["UPLOAD_FOLDER"] = str(UPLOAD_DIR)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024
app.config["WTF_CSRF_TIME_LIMIT"] = None
app.config["RATELIMIT_STORAGE_URI"] = "memory://"

csrf = CSRFProtect(app)
limiter = Limiter(key_func=get_remote_address, app=app, default_limits=[])


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(app.config["DATABASE"])
        g.db.row_factory = sqlite3.Row
    return g.db


@app.context_processor
def inject_helpers():
    return {
        "csrf_token": generate_csrf,
        "lifecycle_statuses": LIFECYCLE_STATUSES,
        "submission_mode_options": SUBMISSION_MODE_OPTIONS,
    }


@app.teardown_appcontext
def close_db(_error):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def table_columns(db, table_name):
    return {row[1] for row in db.execute(f"PRAGMA table_info({table_name})")}


def ensure_column(db, table_name, column_name, definition):
    if table_name not in SCHEMA_WHITELIST or column_name not in SCHEMA_WHITELIST[table_name]:
        raise ValueError(f"Unsafe schema update attempted for {table_name}.{column_name}")
    existing_columns = table_columns(db, table_name)
    if column_name not in existing_columns:
        db.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {definition}")


def migrate_project_approvals_table(db):
    ensure_column(db, "project_approvals", "issue_date", "TEXT NOT NULL DEFAULT ''")
    ensure_column(db, "project_approvals", "expiry_date", "TEXT NOT NULL DEFAULT ''")
    ensure_column(db, "project_approvals", "approval_notes", "TEXT NOT NULL DEFAULT ''")


def migrate_documents_table(db):
    ensure_column(db, "documents", "document_title", "TEXT NOT NULL DEFAULT ''")
    ensure_column(db, "documents", "version_notes", "TEXT NOT NULL DEFAULT ''")


def compliance_table_needs_rebuild(db):
    existing_columns = table_columns(db, "compliance_items")
    required_columns = {
        "approval_type",
        "frequency",
        "schedule_source",
        "submitted_to",
        "submission_mode",
        "responsible_person",
        "acknowledgment_number",
        "remarks",
        "source_type",
        "source_batch_id",
    }
    if not required_columns.issubset(existing_columns):
        return True
    create_sql = db.execute(
        "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'compliance_items'"
    ).fetchone()
    sql_text = (create_sql["sql"] if create_sql else "") or ""
    return "On hold" not in sql_text or "Not applicable" not in sql_text


def migrate_compliance_items_table(db):
    if not compliance_table_needs_rebuild(db):
        return

    db.execute("PRAGMA foreign_keys = OFF")
    db.execute("ALTER TABLE compliance_items RENAME TO compliance_items_legacy")
    db.execute(
        """
        CREATE TABLE compliance_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            approval_type TEXT NOT NULL DEFAULT '',
            source_type TEXT NOT NULL DEFAULT 'manual',
            source_batch_id INTEGER,
            condition_description TEXT NOT NULL,
            action_to_be_taken TEXT NOT NULL DEFAULT '',
            due_date TEXT NOT NULL DEFAULT '',
            frequency TEXT NOT NULL DEFAULT 'General',
            schedule_source TEXT NOT NULL DEFAULT '',
            submitted_to TEXT NOT NULL DEFAULT '',
            submission_mode TEXT NOT NULL DEFAULT '',
            responsible_person TEXT NOT NULL DEFAULT '',
            acknowledgment_number TEXT NOT NULL DEFAULT '',
            remarks TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL CHECK (status IN ('Pending', 'Completed', 'On hold', 'Not applicable')) DEFAULT 'Pending',
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        )
        """
    )

    legacy_columns = table_columns(db, "compliance_items_legacy")

    def select_expr(column_name, fallback="''"):
        return column_name if column_name in legacy_columns else fallback

    db.execute(
        f"""
        INSERT INTO compliance_items (
            id,
            project_id,
            approval_type,
            source_type,
            source_batch_id,
            condition_description,
            action_to_be_taken,
            due_date,
            frequency,
            schedule_source,
            submitted_to,
            submission_mode,
            responsible_person,
            acknowledgment_number,
            remarks,
            status,
            created_at
        )
        SELECT
            id,
            project_id,
            {select_expr('approval_type')},
            {select_expr('source_type', "'manual'")},
            {select_expr('source_batch_id', 'NULL')},
            condition_description,
            {select_expr('action_to_be_taken')},
            {select_expr('due_date')},
            {select_expr('frequency', "'General'")},
            {select_expr('schedule_source')},
            {select_expr('submitted_to')},
            {select_expr('submission_mode')},
            {select_expr('responsible_person')},
            {select_expr('acknowledgment_number')},
            {select_expr('remarks')},
            CASE
                WHEN status IN ('Pending', 'Completed', 'On hold', 'Not applicable') THEN status
                ELSE 'Pending'
            END,
            COALESCE({select_expr('created_at', 'CURRENT_TIMESTAMP')}, CURRENT_TIMESTAMP)
        FROM compliance_items_legacy
        """
    )
    db.execute("DROP TABLE compliance_items_legacy")
    db.execute("PRAGMA foreign_keys = ON")


def create_history_table(db):
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS compliance_item_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            compliance_item_id INTEGER,
            project_id INTEGER NOT NULL,
            change_type TEXT NOT NULL,
            field_label TEXT NOT NULL DEFAULT '',
            previous_value TEXT NOT NULL DEFAULT '',
            new_value TEXT NOT NULL DEFAULT '',
            item_snapshot TEXT NOT NULL DEFAULT '',
            actor_email TEXT NOT NULL DEFAULT '',
            changed_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )


def create_occurrence_table(db):
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS compliance_occurrences (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            compliance_item_id INTEGER NOT NULL,
            occurrence_date TEXT NOT NULL,
            status TEXT NOT NULL CHECK (status IN ('Pending', 'Completed', 'On hold', 'Not applicable')) DEFAULT 'Pending',
            response_notes TEXT NOT NULL DEFAULT '',
            completed_at TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(compliance_item_id, occurrence_date),
            FOREIGN KEY (compliance_item_id) REFERENCES compliance_items(id) ON DELETE CASCADE
        )
        """
    )


def create_compliance_report_tables(db):
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS compliance_report_responses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            approval_type TEXT NOT NULL,
            compliance_item_id INTEGER NOT NULL,
            period_label TEXT NOT NULL,
            period_due_date TEXT NOT NULL DEFAULT '',
            response_text TEXT NOT NULL DEFAULT '',
            attachment_original_filename TEXT NOT NULL DEFAULT '',
            attachment_stored_filename TEXT NOT NULL DEFAULT '',
            annexure_number INTEGER NOT NULL DEFAULT 0,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(project_id, approval_type, compliance_item_id, period_label),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE,
            FOREIGN KEY (compliance_item_id) REFERENCES compliance_items(id) ON DELETE CASCADE
        )
        """
    )


def create_ec_extraction_tables(db):
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS ec_extraction_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            approval_type TEXT NOT NULL DEFAULT 'EC',
            source_filename TEXT NOT NULL DEFAULT '',
            stored_filename TEXT NOT NULL DEFAULT '',
            reference_number TEXT NOT NULL DEFAULT '',
            issue_date TEXT NOT NULL DEFAULT '',
            validity_text TEXT NOT NULL DEFAULT '',
            proponent_name TEXT NOT NULL DEFAULT '',
            location_text TEXT NOT NULL DEFAULT '',
            raw_text TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL DEFAULT 'draft',
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            confirmed_at TEXT NOT NULL DEFAULT '',
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS ec_extraction_batch_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id INTEGER NOT NULL,
            item_order INTEGER NOT NULL DEFAULT 0,
            condition_description TEXT NOT NULL DEFAULT '',
            action_to_be_taken TEXT NOT NULL DEFAULT '',
            is_selected INTEGER NOT NULL DEFAULT 1,
            FOREIGN KEY (batch_id) REFERENCES ec_extraction_batches(id) ON DELETE CASCADE
        )
        """
    )


def init_db():
    INSTANCE_DIR.mkdir(exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    EXTRACTION_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    db = sqlite3.connect(app.config["DATABASE"])
    db.row_factory = sqlite3.Row
    db.executescript(
        """
        PRAGMA foreign_keys = ON;

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            client_name TEXT NOT NULL,
            location TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS project_approvals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            approval_type TEXT NOT NULL,
            issue_date TEXT NOT NULL DEFAULT '',
            expiry_date TEXT NOT NULL DEFAULT '',
            approval_notes TEXT NOT NULL DEFAULT '',
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS compliance_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            approval_type TEXT NOT NULL DEFAULT '',
            source_type TEXT NOT NULL DEFAULT 'manual',
            source_batch_id INTEGER,
            condition_description TEXT NOT NULL,
            action_to_be_taken TEXT NOT NULL DEFAULT '',
            due_date TEXT NOT NULL DEFAULT '',
            frequency TEXT NOT NULL DEFAULT 'General',
            schedule_source TEXT NOT NULL DEFAULT '',
            submitted_to TEXT NOT NULL DEFAULT '',
            submission_mode TEXT NOT NULL DEFAULT '',
            responsible_person TEXT NOT NULL DEFAULT '',
            acknowledgment_number TEXT NOT NULL DEFAULT '',
            remarks TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL CHECK (status IN ('Pending', 'Completed', 'On hold', 'Not applicable')) DEFAULT 'Pending',
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            compliance_item_id INTEGER NOT NULL,
            original_filename TEXT NOT NULL,
            stored_filename TEXT NOT NULL,
            document_title TEXT NOT NULL DEFAULT '',
            version_notes TEXT NOT NULL DEFAULT '',
            uploaded_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (compliance_item_id) REFERENCES compliance_items(id) ON DELETE CASCADE
        );
        """
    )
    migrate_project_approvals_table(db)
    migrate_compliance_items_table(db)
    migrate_documents_table(db)
    create_history_table(db)
    create_occurrence_table(db)
    create_compliance_report_tables(db)
    create_ec_extraction_tables(db)
    db.execute(
        """
        UPDATE compliance_items
        SET approval_type = COALESCE(
            NULLIF(approval_type, ''),
            NULLIF(schedule_source, ''),
            (
                SELECT pa.approval_type
                FROM project_approvals pa
                WHERE pa.project_id = compliance_items.project_id
                ORDER BY CASE pa.approval_type WHEN 'EC' THEN 1 WHEN 'CTE' THEN 2 WHEN 'CTO' THEN 3 ELSE 9 END, pa.id
                LIMIT 1
            ),
            ''
        )
        WHERE approval_type = ''
        """
    )
    db.commit()
    db.close()


def login_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if g.user is None:
            return redirect(url_for("login"))
        return view(**kwargs)

    return wrapped_view


@app.before_request
def load_logged_in_user():
    user_id = session.get("user_id")
    g.user = None
    g.notification_items = []
    g.notification_count = 0
    if user_id is not None:
        g.user = get_db().execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if g.user is not None:
            _, upcoming_items = build_user_reminders(g.user["id"])
            g.notification_items = upcoming_items[:6]
            g.notification_count = len(upcoming_items)


@app.errorhandler(CSRFError)
def handle_csrf_error(_error):
    flash("Your session form token expired or was invalid. Please try that action again.", "error")
    return redirect(request.referrer or url_for("dashboard"))


def current_user_email():
    return g.user["email"] if getattr(g, "user", None) else ""


def get_owned_project(project_id):
    project = get_db().execute(
        "SELECT * FROM projects WHERE id = ? AND user_id = ?", (project_id, g.user["id"])
    ).fetchone()
    if project is None:
        abort(404)
    return project


def get_owned_project_approval(project_id, approval_type):
    project = get_owned_project(project_id)
    approval = get_db().execute(
        """
        SELECT *
        FROM project_approvals
        WHERE project_id = ? AND approval_type = ?
        """,
        (project_id, approval_type),
    ).fetchone()
    if approval is None:
        abort(404)
    return project, approval


def get_owned_ec_extraction_batch(batch_id):
    batch = get_db().execute(
        """
        SELECT b.*, p.user_id, p.name AS project_name
        FROM ec_extraction_batches b
        JOIN projects p ON p.id = b.project_id
        WHERE b.id = ? AND p.user_id = ?
        """,
        (batch_id, g.user["id"]),
    ).fetchone()
    if batch is None:
        abort(404)
    return batch


def get_owned_compliance_item(item_id):
    item = get_db().execute(
        """
        SELECT ci.*, p.user_id, p.name AS project_name
        FROM compliance_items ci
        JOIN projects p ON p.id = ci.project_id
        WHERE ci.id = ? AND p.user_id = ?
        """,
        (item_id, g.user["id"]),
    ).fetchone()
    if item is None:
        abort(404)
    return item


def allowed_file(filename):
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


def allowed_import_file(filename):
    return Path(filename).suffix.lower() in ALLOWED_IMPORT_EXTENSIONS


def normalize_header(value):
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def clean_text(value):
    return str(value or "").replace("\xa0", " ").strip()


def normalize_status(value):
    cleaned = normalize_header(value)
    mapping = {
        "pending": "Pending",
        "completed": "Completed",
        "on hold": "On hold",
        "hold": "On hold",
        "not applicable": "Not applicable",
        "na": "Not applicable",
        "n/a": "Not applicable",
    }
    return mapping.get(cleaned, "Pending")


def normalize_due_date(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return ""

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def normalize_frequency(value, due_date=""):
    cleaned = normalize_header(value)
    frequency_map = {
        "general": "General",
        "ongoing": "General",
        "one-time": "One-time",
        "one time": "One-time",
        "onetime": "One-time",
        "monthly": "Monthly",
        "quarterly": "Quarterly",
        "half yearly": "Half-yearly",
        "half-yearly": "Half-yearly",
        "half yearly basis": "Half-yearly",
        "yearly": "Yearly",
        "annual": "Yearly",
    }
    if cleaned in frequency_map:
        return frequency_map[cleaned]
    return "One-time" if due_date else "General"


def normalize_approval_type(value, allowed_options=None):
    cleaned = clean_text(value).upper()
    if not cleaned:
        return ""
    options = allowed_options or DEFAULT_APPROVAL_TYPES
    normalized_lookup = {option.upper(): option for option in options}
    return normalized_lookup.get(cleaned, "")


def normalize_submission_mode(value):
    cleaned = clean_text(value)
    if not cleaned:
        return ""
    for option in SUBMISSION_MODE_OPTIONS:
        if cleaned.lower() == option.lower():
            return option
    return cleaned


def parse_iso_date(value):
    cleaned = clean_text(value)
    if not cleaned:
        return None
    try:
        return datetime.strptime(cleaned, "%Y-%m-%d").date()
    except ValueError:
        return None


def add_months(source_date, months):
    month = source_date.month - 1 + months
    year = source_date.year + month // 12
    month = month % 12 + 1
    day = min(source_date.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def recurrence_interval_months(frequency):
    return {
        "Monthly": 1,
        "Quarterly": 3,
        "Half-yearly": 6,
        "Yearly": 12,
    }.get(frequency, 0)


def is_recurring_frequency(frequency):
    return frequency in {"Monthly", "Quarterly", "Half-yearly", "Yearly"}


def looks_like_section_header(text):
    cleaned = clean_text(text)
    if not cleaned:
        return False
    uppercase_ratio = sum(1 for char in cleaned if char.isupper()) / max(
        1, sum(1 for char in cleaned if char.isalpha())
    )
    return cleaned.startswith(("PART ", "SECTION ", "CHAPTER ")) or cleaned.endswith(":") or uppercase_ratio > 0.75


def normalize_pdf_text(text):
    cleaned = (text or "").replace("\xa0", " ").replace("\uf0b7", " ")
    cleaned = cleaned.replace("\r", "\n")
    lines = [" ".join(line.split()) for line in cleaned.splitlines()]
    return "\n".join(line for line in lines if line.strip())


def extract_pdf_text(file_path):
    reader = PdfReader(str(file_path))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def parse_issue_date_value(raw_text):
    if not raw_text:
        return ""
    cleaned = raw_text.replace("st", "").replace("nd", "").replace("rd", "").replace("th", "")
    return normalize_due_date(cleaned)


def extract_ec_metadata(raw_text):
    metadata = {
        "reference_number": "",
        "issue_date": "",
        "validity_text": "",
        "proponent_name": "",
        "location_text": "",
    }
    text = raw_text
    patterns = {
        "reference_number": [
            r"EC\s+letter\s+no\.?\s*[:\-]\s*(.+?)(?:;|\n|Date\s*:)",
            r"reference\s+no\.?\s*[:\-]\s*(.+?)(?:\n|Date\s*:)",
        ],
        "issue_date": [
            r"Date\s*[:\-]\s*([^\n;]+)",
            r"issue\s+date\s*[:\-]\s*([^\n;]+)",
        ],
        "validity_text": [
            r"(?:validity|valid up to|shall be valid[^.\n]*)\s*[:\-]?\s*([^\n]+)",
        ],
        "proponent_name": [
            r"Project\s+Proponent\s*[:\-]?\s*([^\n]+)",
            r"Proponent\s+Name\s*[:\-]?\s*([^\n]+)",
        ],
        "location_text": [
            r"Project\s+Address\s*[:\-]?\s*([^\n]+(?:\n[^\n]+)?)",
            r"Location\s*[:\-]?\s*([^\n]+)",
        ],
    }
    for key, regex_list in patterns.items():
        for pattern in regex_list:
            match = re.search(pattern, text, flags=re.IGNORECASE)
            if match:
                metadata[key] = clean_text(match.group(1))
                break
    metadata["issue_date"] = parse_issue_date_value(metadata["issue_date"])
    return metadata


def extract_condition_candidates(raw_text):
    text = normalize_pdf_text(raw_text)
    condition_blocks = []
    for match in re.finditer(r"(?m)(?:^|\n)\s*(\d{1,3})\.\s+(.+?)(?=(?:\n\s*\d{1,3}\.\s)|\Z)", text, flags=re.DOTALL):
        block = clean_text(match.group(2))
        if not block:
            continue
        block = re.sub(r"\s+", " ", block).strip()
        if len(block) < 20:
            continue
        condition_blocks.append(block)

    if condition_blocks:
        deduped = []
        seen = set()
        for block in condition_blocks:
            normalized = block.lower()
            if normalized in seen:
                continue
            seen.add(normalized)
            deduped.append(block)
        return deduped

    sentences = [clean_text(part) for part in re.split(r"\n+", text) if clean_text(part)]
    return [sentence for sentence in sentences if len(sentence) > 40][:40]


def build_default_action_from_condition(condition_text):
    summary = " ".join(condition_text.split()[:18]).strip()
    if summary:
        return f"Review this EC condition and define the evidence / submission steps for: {summary}..."
    return ""


def parse_import_matrix(rows):
    if not rows:
        return []

    first_row = [clean_text(cell) for cell in rows[0]]
    mapped_headers = [IMPORT_HEADER_ALIASES.get(normalize_header(cell)) for cell in first_row]
    is_structured = any(mapped_headers)
    data_rows = rows[1:] if is_structured else rows
    normalized_rows = []

    for row in data_rows:
        row = tuple(row)
        if is_structured:
            mapped = {}
            for index, mapped_key in enumerate(mapped_headers):
                if mapped_key:
                    mapped[mapped_key] = row[index] if index < len(row) else None

            condition_description = clean_text(mapped.get("condition_description"))
            action_to_be_taken = clean_text(mapped.get("action_to_be_taken"))
            due_date = normalize_due_date(mapped.get("due_date"))
            status_raw = clean_text(mapped.get("status"))
            frequency_raw = clean_text(mapped.get("frequency"))
            serial_number = clean_text(mapped.get("serial_number"))

            if not condition_description:
                continue
            if (
                not serial_number
                and not action_to_be_taken
                and not due_date
                and not status_raw
                and not frequency_raw
                and looks_like_section_header(condition_description)
            ):
                continue

            schedule_source = clean_text(mapped.get("schedule_source"))
            submitted_to = clean_text(mapped.get("submitted_to"))
            submission_mode = normalize_submission_mode(mapped.get("submission_mode"))
            responsible_person = clean_text(mapped.get("responsible_person"))
            acknowledgment_number = clean_text(mapped.get("acknowledgment_number"))
            remarks = clean_text(mapped.get("remarks"))
            approval_type = clean_text(mapped.get("approval_type"))
        else:
            non_empty_cells = [clean_text(cell) for cell in row if clean_text(cell)]
            if not non_empty_cells:
                continue
            if IMPORT_HEADER_ALIASES.get(normalize_header(non_empty_cells[0])):
                continue

            condition_description = non_empty_cells[0]
            action_to_be_taken = non_empty_cells[1] if len(non_empty_cells) > 1 else ""
            due_date = normalize_due_date(non_empty_cells[2]) if len(non_empty_cells) > 2 else ""
            status_raw = non_empty_cells[3] if len(non_empty_cells) > 3 else ""
            frequency_raw = non_empty_cells[4] if len(non_empty_cells) > 4 else ""
            schedule_source = ""
            submitted_to = ""
            submission_mode = ""
            responsible_person = ""
            acknowledgment_number = ""
            remarks = ""
            approval_type = ""

        normalized_rows.append(
            {
                "approval_type": approval_type,
                "condition_description": condition_description,
                "action_to_be_taken": action_to_be_taken,
                "due_date": due_date,
                "status": normalize_status(status_raw),
                "frequency": normalize_frequency(frequency_raw, due_date),
                "schedule_source": schedule_source,
                "submitted_to": submitted_to,
                "submission_mode": submission_mode,
                "responsible_person": responsible_person,
                "acknowledgment_number": acknowledgment_number,
                "remarks": remarks,
            }
        )

    return normalized_rows


def parse_import_rows(file_storage):
    extension = Path(file_storage.filename).suffix.lower()

    if extension == ".csv":
        content = file_storage.stream.read().decode("utf-8-sig", errors="ignore")
        reader = csv.reader(StringIO(content))
        rows = list(reader)
    else:
        workbook = load_workbook(file_storage, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        rows = values

    return parse_import_matrix(rows)


def build_sample_import_csv():
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "Approval Type",
            "Condition Description",
            "Action To Be Taken",
            "Due Date",
            "Status",
            "Frequency",
            "Schedule Source",
            "Submitted To",
            "Submission Mode",
            "Responsible Person",
            "Acknowledgment Number",
            "Remarks",
        ]
    )
    writer.writerow(
        [
            "EC",
            "Submit half-yearly compliance report to the regional office.",
            "Prepare the report, review internally, and submit before the deadline.",
            "2026-07-15",
            "Pending",
            "One-time",
            "",
            "Regional Office",
            "Portal",
            "Consultant lead",
            "EC-HY-2026-01",
            "Attach the signed copy after submission.",
        ]
    )
    writer.writerow(
        [
            "EC",
            "Submit quarterly monitoring summary from EC grant date.",
            "Compile monitoring results and submit to the authority.",
            "",
            "Pending",
            "Quarterly",
            "EC",
            "State authority",
            "Email",
            "Site environment manager",
            "",
            "",
        ]
    )
    writer.writerow(
        [
            "Display the latest EC copy at the site office and entry gate.",
            "",
            "",
            "Pending",
            "General",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    return output.getvalue()


def build_approval_form_entries(selected_rows=None):
    selected_map = {}
    if selected_rows:
        for row in selected_rows:
            approval_type = row["approval_type"] if isinstance(row, sqlite3.Row) else row["approval_type"]
            selected_map[approval_type] = {
                "issue_date": row["issue_date"],
                "expiry_date": row["expiry_date"],
                "approval_notes": row.get("approval_notes", "") if isinstance(row, dict) else row["approval_notes"],
            }

    return [
        {
            "approval_type": approval_type,
            "selected": approval_type in selected_map,
            "issue_date": selected_map.get(approval_type, {}).get("issue_date", ""),
            "expiry_date": selected_map.get(approval_type, {}).get("expiry_date", ""),
            "approval_notes": selected_map.get(approval_type, {}).get("approval_notes", ""),
        }
        for approval_type in DEFAULT_APPROVAL_TYPES
    ]


def parse_approval_form(form):
    selected_entries = []
    form_entries = []
    for approval_type in DEFAULT_APPROVAL_TYPES:
        selected = bool(form.get(f"approval_enabled_{approval_type}"))
        issue_date = form.get(f"issue_date_{approval_type}", "").strip()
        expiry_date = form.get(f"expiry_date_{approval_type}", "").strip()
        approval_notes = form.get(f"approval_notes_{approval_type}", "").strip()
        form_entries.append(
            {
                "approval_type": approval_type,
                "selected": selected,
                "issue_date": issue_date,
                "expiry_date": expiry_date,
                "approval_notes": approval_notes,
            }
        )
        if selected:
            selected_entries.append(
                {
                    "approval_type": approval_type,
                    "issue_date": issue_date,
                    "expiry_date": expiry_date,
                    "approval_notes": approval_notes,
                }
            )
    return selected_entries, form_entries


def derive_item_status(status, due_date_value):
    if status in {"Completed", "On hold", "Not applicable"}:
        return status
    if not due_date_value:
        return "Pending"

    today = date.today().isoformat()
    seven_days = (date.today() + timedelta(days=7)).isoformat()
    if due_date_value < today:
        return "Overdue"
    if due_date_value <= seven_days:
        return "Due in 7 days"
    return "Pending"


def annotate_compliance_items(rows, approval_lookup):
    occurrence_map = fetch_occurrence_rows([row["id"] for row in rows])
    annotated = []
    for row in rows:
        item = dict(row)
        item_occurrences = build_occurrence_records(item, approval_lookup, occurrence_map.get(item["id"], {}))
        item["occurrences"] = item_occurrences
        item["occurrence_summary"] = summarize_occurrences(item_occurrences)
        item["is_recurring"] = bool(item_occurrences)
        item["next_due_date"] = (
            item["occurrence_summary"]["next_due_date"] if item_occurrences else compute_next_due_date(item, approval_lookup)
        )
        item["display_due_date"] = item["next_due_date"] or item["due_date"]
        item["derived_status"] = derive_item_status(item["status"], item["display_due_date"])
        if item_occurrences:
            if item["occurrence_summary"]["overdue"]:
                item["derived_status"] = "Overdue"
            elif item["occurrence_summary"]["upcoming"]:
                item["derived_status"] = "Due in 7 days"
            elif item["occurrence_summary"]["pending"]:
                item["derived_status"] = "Pending"
            elif item["occurrence_summary"]["completed"] == len(item_occurrences):
                item["derived_status"] = "Completed"
        item["has_due_date"] = bool(item["display_due_date"])
        item["is_overdue"] = item["derived_status"] == "Overdue"
        item["is_due_soon"] = item["derived_status"] == "Due in 7 days"
        item["schedule_label"] = item["frequency"]
        if item.get("schedule_source"):
            item["schedule_label"] = f"{item['frequency']} from {item['schedule_source']}"
        item["status_summary_label"] = format_occurrence_summary(item["occurrence_summary"]) if item_occurrences else item["derived_status"]
        annotated.append(item)
    return annotated


def build_status_counts(items):
    counts = {filter_name: 0 for filter_name in STATUS_FILTERS}
    counts["All"] = len(items)
    for item in items:
        counts[item["derived_status"]] += 1
    return counts


def build_frequency_counts(items):
    counts = {"All": len(items)}
    for frequency in FREQUENCY_TYPES:
        counts[frequency] = sum(1 for item in items if item["frequency"] == frequency)
    return counts


def filter_compliance_items(items, status_filter, frequency_filter):
    filtered = items
    if status_filter != "All":
        filtered = [item for item in filtered if item["derived_status"] == status_filter]
    if frequency_filter != "All":
        filtered = [item for item in filtered if item["frequency"] == frequency_filter]
    return filtered


def annotate_approval_rows(rows):
    today = date.today().isoformat()
    warning_cutoff = (date.today() + timedelta(days=30)).isoformat()
    annotated = []
    for row in rows:
        approval = dict(row)
        approval["expiry_state"] = "No expiry set"
        if approval["expiry_date"]:
            if approval["expiry_date"] < today:
                approval["expiry_state"] = "Expired"
            elif approval["expiry_date"] <= warning_cutoff:
                approval["expiry_state"] = "Expiring soon"
            else:
                approval["expiry_state"] = "Active"
        annotated.append(approval)
    return annotated


def build_approval_lookup(approval_rows):
    return {row["approval_type"]: dict(row) for row in approval_rows}


def resolve_schedule_anchor(item, approval_lookup):
    schedule_source = clean_text(item.get("schedule_source"))
    if schedule_source and schedule_source in approval_lookup:
        approval = approval_lookup[schedule_source]
        return parse_iso_date(approval.get("issue_date")), parse_iso_date(approval.get("expiry_date"))
    return parse_iso_date(item.get("due_date")), None


def compute_next_due_date(item, approval_lookup, from_date=None):
    if item["frequency"] in {"General", ""}:
        return item.get("due_date", "")

    from_date = from_date or date.today()
    interval_months = recurrence_interval_months(item["frequency"])
    if not interval_months:
        return item.get("due_date", "")

    anchor_date, end_date = resolve_schedule_anchor(item, approval_lookup)
    if not anchor_date:
        return item.get("due_date", "")

    occurrence = anchor_date
    last_occurrence = None
    while occurrence <= from_date:
        last_occurrence = occurrence
        occurrence = add_months(occurrence, interval_months)

    if item.get("status") == "Pending" and last_occurrence is not None:
        candidate = last_occurrence
    else:
        candidate = occurrence

    if end_date and candidate > end_date:
        return ""
    return candidate.isoformat()


def generate_schedule_dates(item, approval_lookup, window_start, window_end):
    if item["frequency"] in {"General", ""}:
        due_date = parse_iso_date(item.get("due_date"))
        if due_date and window_start <= due_date <= window_end:
            return [due_date]
        return []

    interval_months = recurrence_interval_months(item["frequency"])
    if not interval_months:
        due_date = parse_iso_date(item.get("due_date"))
        if due_date and window_start <= due_date <= window_end:
            return [due_date]
        return []

    anchor_date, end_date = resolve_schedule_anchor(item, approval_lookup)
    if not anchor_date:
        return []

    occurrence = anchor_date
    dates = []
    while occurrence <= window_end:
        if end_date and occurrence > end_date:
            break
        if occurrence >= window_start:
            dates.append(occurrence)
        occurrence = add_months(occurrence, interval_months)
    return dates


def build_schedule_preview(frequency, due_date, schedule_source, approval_lookup, occurrences=4):
    if frequency == "General":
        return "No recurring schedule. Use this for general site conditions without a deadline."

    if frequency == "One-time":
        if due_date:
            return f"One-time deadline on {due_date}."
        return "One-time item with no date selected yet."

    interval_months = recurrence_interval_months(frequency)
    if not interval_months:
        return "Schedule preview becomes available after selecting a valid frequency."

    item = {
        "frequency": frequency,
        "due_date": due_date,
        "schedule_source": schedule_source,
    }
    anchor_date, end_date = resolve_schedule_anchor(item, approval_lookup)
    if not anchor_date:
        anchor_label = schedule_source or "a starting date"
        return f"Select {anchor_label} issue date or a custom date to preview the recurring schedule."

    dates = []
    occurrence = anchor_date
    while len(dates) < occurrences:
        if end_date and occurrence > end_date:
            break
        dates.append(occurrence.isoformat())
        occurrence = add_months(occurrence, interval_months)

    if not dates:
        return "No recurring dates fall inside the selected approval validity window."

    suffix = f" until {end_date.isoformat()}" if end_date else ""
    return f"Upcoming schedule: {', '.join(dates)}{suffix}."


def build_project_calendar(project_id, months=1):
    project = get_owned_project(project_id)
    approvals = fetch_project_approvals(project_id)
    approval_lookup = build_approval_lookup(approvals)
    due_items = fetch_project_compliance_items(project_id)

    event_map = {}

    def add_event(event_date, label, kind, href):
        if not event_date:
            return
        event_map.setdefault(event_date, []).append({"label": label, "kind": kind, "href": href})

    start = date.today().replace(day=1)
    current_year = start.year
    current_month = start.month
    months_data = []
    window_end = add_months(start, months) - timedelta(days=1)

    for item in due_items:
        compact_label = item["condition_description"][:56].rstrip()
        if len(item["condition_description"]) > 56:
            compact_label += "..."
        for occurrence in generate_schedule_dates(dict(item), approval_lookup, start, window_end):
            occurrence_kind = "green"
            if occurrence.isoformat() < date.today().isoformat():
                occurrence_kind = "red"
            elif occurrence.isoformat() <= (date.today() + timedelta(days=15)).isoformat():
                occurrence_kind = "amber"
            add_event(
                occurrence.isoformat(),
                f"{item['frequency']}: {compact_label}",
                occurrence_kind,
                url_for("approval_detail", project_id=project["id"], approval_type=item["approval_type"]),
            )

    for approval in approvals:
        add_event(
            approval["issue_date"],
            f"{approval['approval_type']} issue date",
            "green",
            url_for("approval_detail", project_id=project["id"], approval_type=approval["approval_type"]),
        )
        expiry_kind = "green"
        if approval["expiry_date"]:
            if approval["expiry_date"] < date.today().isoformat():
                expiry_kind = "red"
            elif approval["expiry_date"] <= (date.today() + timedelta(days=15)).isoformat():
                expiry_kind = "amber"
        add_event(
            approval["expiry_date"],
            f"{approval['approval_type']} expiry date",
            expiry_kind,
            url_for("approval_detail", project_id=project["id"], approval_type=approval["approval_type"]),
        )

    for _ in range(months):
        month_matrix = calendar.monthcalendar(current_year, current_month)
        days = []
        for week in month_matrix:
            for day_number in week:
                if day_number == 0:
                    days.append({"day": "", "date": "", "events": [], "is_today": False})
                    continue

                current_date = date(current_year, current_month, day_number)
                iso_date = current_date.isoformat()
                days.append(
                    {
                        "day": day_number,
                        "date": iso_date,
                        "events": event_map.get(iso_date, []),
                        "is_today": current_date == date.today(),
                    }
                )

        months_data.append({"label": f"{calendar.month_name[current_month]} {current_year}", "days": days})

        if current_month == 12:
            current_month = 1
            current_year += 1
        else:
            current_month += 1

    return months_data


def fetch_project_approvals(project_id):
    return get_db().execute(
        """
        SELECT approval_type, issue_date, expiry_date, approval_notes
        FROM project_approvals
        WHERE project_id = ?
        ORDER BY approval_type
        """,
        (project_id,),
    ).fetchall()


def fetch_project_compliance_items(project_id, approval_type=None):
    filters = ["ci.project_id = ?"]
    params = [project_id]
    if approval_type:
        filters.append("ci.approval_type = ?")
        params.append(approval_type)
    return get_db().execute(
        f"""
        SELECT ci.*,
               COUNT(d.id) AS document_count
        FROM compliance_items ci
        LEFT JOIN documents d ON d.compliance_item_id = ci.id
        WHERE {' AND '.join(filters)}
        GROUP BY ci.id
        ORDER BY CASE WHEN ci.due_date = '' THEN 1 ELSE 0 END, ci.due_date ASC, ci.created_at DESC
        """,
        params,
    ).fetchall()


def fetch_project_documents(project_id, approval_type=None):
    filters = ["ci.project_id = ?"]
    params = [project_id]
    if approval_type:
        filters.append("ci.approval_type = ?")
        params.append(approval_type)
    return get_db().execute(
        f"""
        SELECT d.*, ci.project_id, ci.approval_type, ci.condition_description
        FROM documents d
        JOIN compliance_items ci ON ci.id = d.compliance_item_id
        WHERE {' AND '.join(filters)}
        ORDER BY d.uploaded_at DESC
        """,
        params,
    ).fetchall()


def fetch_project_history(project_id, limit=12, approval_type=None):
    params = [project_id]
    filter_sql = ""
    if approval_type:
        filter_sql = """
            AND compliance_item_id IN (
                SELECT id FROM compliance_items WHERE project_id = ? AND approval_type = ?
            )
        """
        params.extend([project_id, approval_type])
    params.append(limit)
    return get_db().execute(
        f"""
        SELECT *
        FROM compliance_item_history
        WHERE project_id = ?
        {filter_sql}
        ORDER BY changed_at DESC, id DESC
        LIMIT ?
        """,
        params,
    ).fetchall()


def fetch_extraction_batch_items(batch_id):
    return get_db().execute(
        """
        SELECT *
        FROM ec_extraction_batch_items
        WHERE batch_id = ?
        ORDER BY item_order ASC, id ASC
        """,
        (batch_id,),
    ).fetchall()


def build_documents_by_item(documents):
    docs_by_item = {}
    for document in documents:
        docs_by_item.setdefault(document["compliance_item_id"], []).append(document)
    return docs_by_item


def fetch_occurrence_rows(item_ids):
    if not item_ids:
        return {}
    placeholders = ",".join("?" for _ in item_ids)
    rows = get_db().execute(
        f"""
        SELECT *
        FROM compliance_occurrences
        WHERE compliance_item_id IN ({placeholders})
        ORDER BY occurrence_date ASC
        """,
        item_ids,
    ).fetchall()
    occurrence_map = {}
    for row in rows:
        occurrence_map.setdefault(row["compliance_item_id"], {})[row["occurrence_date"]] = dict(row)
    return occurrence_map


def build_occurrence_records(item, approval_lookup, occurrence_rows, window_limit=60):
    if item["frequency"] in {"General", "One-time", ""}:
        return []

    interval_months = recurrence_interval_months(item["frequency"])
    if not interval_months:
        return []

    anchor_date, end_date = resolve_schedule_anchor(item, approval_lookup)
    if not anchor_date:
        return []

    records = []
    current_date = anchor_date
    count = 0
    while count < window_limit:
        if end_date and current_date > end_date:
            break
        iso_date = current_date.isoformat()
        stored = occurrence_rows.get(iso_date, {})
        stored_status = stored.get("status") or "Pending"
        derived_status = derive_item_status(stored_status, iso_date)
        records.append(
            {
                "date": iso_date,
                "label": current_date.strftime("%b %Y"),
                "status": stored_status,
                "derived_status": derived_status,
                "response_notes": stored.get("response_notes", ""),
            }
        )
        current_date = add_months(current_date, interval_months)
        count += 1
    return records


def summarize_occurrences(occurrences):
    if not occurrences:
        return {"completed": 0, "overdue": 0, "upcoming": 0, "pending": 0, "next_due_date": ""}

    summary = {"completed": 0, "overdue": 0, "upcoming": 0, "pending": 0}
    next_due_date = ""
    for occurrence in occurrences:
        if occurrence["status"] == "Completed":
            summary["completed"] += 1
            continue
        if occurrence["derived_status"] == "Overdue":
            summary["overdue"] += 1
        elif occurrence["derived_status"] == "Due in 7 days":
            summary["upcoming"] += 1
        else:
            summary["pending"] += 1
        if not next_due_date:
            next_due_date = occurrence["date"]
    summary["next_due_date"] = next_due_date
    return summary


def format_occurrence_summary(summary):
    parts = []
    if summary["overdue"]:
        parts.append(f"{summary['overdue']} overdue")
    if summary["completed"]:
        parts.append(f"{summary['completed']} completed")
    upcoming_total = summary["upcoming"] + summary["pending"]
    if upcoming_total:
        parts.append(f"{upcoming_total} upcoming")
    return ", ".join(parts) or "No occurrences yet"


def fetch_user_approval_rows(user_id):
    return get_db().execute(
        """
        SELECT pa.project_id, pa.approval_type, pa.issue_date, pa.expiry_date, pa.approval_notes
        FROM project_approvals pa
        JOIN projects p ON p.id = pa.project_id
        WHERE p.user_id = ?
        """,
        (user_id,),
    ).fetchall()


def fetch_user_compliance_rows(user_id):
    return get_db().execute(
        """
        SELECT ci.*, p.name AS project_name, p.location AS project_location, p.client_name
        FROM compliance_items ci
        JOIN projects p ON p.id = ci.project_id
        WHERE p.user_id = ?
        ORDER BY p.name ASC, ci.created_at DESC
        """,
        (user_id,),
    ).fetchall()


def build_project_approval_lookup(approval_rows):
    project_lookup = {}
    for row in approval_rows:
        project_lookup.setdefault(row["project_id"], {})[row["approval_type"]] = dict(row)
    return project_lookup


def annotate_user_compliance_items(user_id):
    approval_lookup = build_project_approval_lookup(fetch_user_approval_rows(user_id))
    rows = fetch_user_compliance_rows(user_id)
    occurrence_map = fetch_occurrence_rows([row["id"] for row in rows])
    annotated = []
    for row in rows:
        item = dict(row)
        project_approvals = approval_lookup.get(item["project_id"], {})
        item_occurrences = build_occurrence_records(item, project_approvals, occurrence_map.get(item["id"], {}))
        item["occurrences"] = item_occurrences
        item["occurrence_summary"] = summarize_occurrences(item_occurrences)
        item["is_recurring"] = bool(item_occurrences)
        item["next_due_date"] = (
            item["occurrence_summary"]["next_due_date"] if item_occurrences else compute_next_due_date(item, project_approvals)
        )
        item["display_due_date"] = item["next_due_date"] or item["due_date"]
        item["derived_status"] = derive_item_status(item["status"], item["display_due_date"])
        if item_occurrences:
            if item["occurrence_summary"]["overdue"]:
                item["derived_status"] = "Overdue"
            elif item["occurrence_summary"]["upcoming"]:
                item["derived_status"] = "Due in 7 days"
            elif item["occurrence_summary"]["pending"]:
                item["derived_status"] = "Pending"
            elif item["occurrence_summary"]["completed"] == len(item_occurrences):
                item["derived_status"] = "Completed"
        annotated.append(item)
    return annotated


def build_user_reminders(user_id, days=7):
    cutoff = (date.today() + timedelta(days=days)).isoformat()
    items = annotate_user_compliance_items(user_id)
    overdue_items = [
        item for item in items if item["status"] == "Pending" and item["derived_status"] == "Overdue"
    ]
    upcoming_items = [
        item for item in items if item["status"] == "Pending" and item["derived_status"] == "Due in 7 days"
    ]
    upcoming_items = [item for item in upcoming_items if item["display_due_date"] <= cutoff]
    overdue_items.sort(key=lambda item: item["display_due_date"])
    upcoming_items.sort(key=lambda item: item["display_due_date"])
    return overdue_items, upcoming_items


def build_project_reminders(items):
    overdue_items = [
        {"id": item["id"], "condition_description": item["condition_description"], "due_date": item["display_due_date"]}
        for item in items
        if item["status"] == "Pending" and item["derived_status"] == "Overdue"
    ]
    upcoming_items = [
        {"id": item["id"], "condition_description": item["condition_description"], "due_date": item["display_due_date"]}
        for item in items
        if item["status"] == "Pending" and item["derived_status"] == "Due in 7 days"
    ]
    return overdue_items, upcoming_items


def fetch_user_documents(user_id):
    action_documents = [
        dict(row)
        for row in get_db().execute(
        """
        SELECT d.*, ci.project_id, p.name AS project_name, p.location AS project_location
        FROM documents d
        JOIN compliance_items ci ON ci.id = d.compliance_item_id
        JOIN projects p ON p.id = ci.project_id
        WHERE p.user_id = ?
        ORDER BY d.uploaded_at DESC
        """,
        (user_id,),
    ).fetchall()
    ]
    for document in action_documents:
        document["source"] = "Action to be taken"

    report_documents = [
        dict(row)
        for row in get_db().execute(
            """
            SELECT crr.id, crr.attachment_original_filename AS original_filename,
                   crr.attachment_stored_filename AS stored_filename,
                   crr.updated_at AS uploaded_at,
                   crr.annexure_number,
                   crr.period_label,
                   p.name AS project_name,
                   p.location AS project_location,
                   crr.approval_type
            FROM compliance_report_responses crr
            JOIN projects p ON p.id = crr.project_id
            WHERE p.user_id = ? AND crr.attachment_stored_filename != ''
            ORDER BY crr.updated_at DESC
            """,
            (user_id,),
        ).fetchall()
    ]
    for document in report_documents:
        document["annexure_label"] = format_annexure_label(
            document["annexure_number"], document.get("period_due_date", ""), document["period_label"]
        )
        document["document_title"] = f"{document['annexure_label']} · {document['period_label']}"
        document["source"] = "Compliance report"

    return sorted(action_documents + report_documents, key=lambda row: row["uploaded_at"], reverse=True)


def build_project_cards(user_id, approval_filter=None):
    project_rows = get_db().execute(
        """
        SELECT p.id, p.name, p.client_name, p.location,
               COUNT(ci.id) AS compliance_count
        FROM projects p
        LEFT JOIN compliance_items ci ON ci.project_id = p.id
        WHERE p.user_id = ?
        GROUP BY p.id
        ORDER BY p.created_at DESC
        """,
        (user_id,),
    ).fetchall()
    approvals_by_project = {}
    for row in fetch_user_approval_rows(user_id):
        approvals_by_project.setdefault(row["project_id"], []).append(row["approval_type"])

    items_by_project = {}
    for item in annotate_user_compliance_items(user_id):
        items_by_project.setdefault(item["project_id"], []).append(item)

    cards = []
    for row in project_rows:
        approval_types = approvals_by_project.get(row["id"], [])
        if approval_filter and approval_filter not in approval_types:
            continue

        project_items = items_by_project.get(row["id"], [])
        overdue_count = sum(1 for item in project_items if item["derived_status"] == "Overdue")
        due_soon_count = sum(1 for item in project_items if item["derived_status"] == "Due in 7 days")
        cards.append(
            {
                "id": row["id"],
                "name": row["name"],
                "client_name": row["client_name"],
                "location": row["location"],
                "compliance_count": row["compliance_count"],
                "approval_types": approval_types,
                "overdue_count": overdue_count,
                "due_soon_count": due_soon_count,
                "all_clear": row["compliance_count"] > 0 and overdue_count == 0 and due_soon_count == 0,
            }
        )
    return cards


def build_project_approval_summaries(project_id):
    approvals = annotate_approval_rows(fetch_project_approvals(project_id))
    approval_lookup = build_approval_lookup(approvals)
    items = annotate_compliance_items(fetch_project_compliance_items(project_id), approval_lookup)
    items_by_approval = {}
    for item in items:
        items_by_approval.setdefault(item["approval_type"], []).append(item)

    summaries = []
    for approval in approvals:
        approval_items = items_by_approval.get(approval["approval_type"], [])
        overdue_count = sum(1 for item in approval_items if item["derived_status"] == "Overdue")
        due_soon_count = sum(1 for item in approval_items if item["derived_status"] == "Due in 7 days")
        summaries.append(
            {
                **approval,
                "condition_count": len(approval_items),
                "overdue_count": overdue_count,
                "due_soon_count": due_soon_count,
            }
        )
    return summaries


def build_approval_portfolio(user_id, approval_type):
    projects_rows = get_db().execute(
        """
        SELECT p.id, p.name, p.location, p.client_name
        FROM projects p
        WHERE p.user_id = ?
        ORDER BY p.name ASC
        """,
        (user_id,),
    ).fetchall()
    approval_rows = [row for row in fetch_user_approval_rows(user_id) if row["approval_type"] == approval_type]
    approval_by_project = {row["project_id"]: annotate_approval_rows([row])[0] for row in approval_rows}
    items = [item for item in annotate_user_compliance_items(user_id) if item["approval_type"] == approval_type]
    items_by_project = {}
    for item in items:
        items_by_project.setdefault(item["project_id"], []).append(item)

    portfolio_rows = []
    for project in projects_rows:
        approval = approval_by_project.get(project["id"])
        project_items = items_by_project.get(project["id"], [])
        overdue_count = sum(1 for item in project_items if item["derived_status"] == "Overdue")
        due_soon_count = sum(1 for item in project_items if item["derived_status"] == "Due in 7 days")
        row = {
            "project_id": project["id"],
            "project_name": project["name"],
            "project_location": project["location"],
            "approval": approval,
            "condition_count": len(project_items),
            "overdue_count": overdue_count,
            "due_soon_count": due_soon_count,
        }
        if approval is None:
            row["portfolio_state"] = "no-record"
        elif approval["expiry_state"] in {"Expired", "Expiring soon"} or overdue_count or due_soon_count:
            row["portfolio_state"] = "needs-attention"
        else:
            row["portfolio_state"] = "healthy"
        portfolio_rows.append(row)

    return {
        "rows": portfolio_rows,
        "needs_attention": [row for row in portfolio_rows if row["portfolio_state"] == "needs-attention"],
        "healthy": [row for row in portfolio_rows if row["portfolio_state"] == "healthy"],
        "no_record": [row for row in portfolio_rows if row["portfolio_state"] == "no-record"],
        "total_projects": len(projects_rows),
        "approval_projects": len(approval_rows),
        "valid_count": sum(1 for row in portfolio_rows if row["approval"] and row["approval"]["expiry_state"] == "Active"),
        "expiring_count": sum(1 for row in portfolio_rows if row["approval"] and row["approval"]["expiry_state"] in {"Expired", "Expiring soon"}),
        "overdue_count": sum(row["overdue_count"] for row in portfolio_rows),
    }


def group_items_for_overview(items, kind):
    grouped = {}
    for item in items:
        project = grouped.setdefault(
            item["project_id"],
            {
                "project_id": item["project_id"],
                "project_name": item["project_name"],
                "project_location": item.get("project_location", ""),
                "items": [],
                "overdue_count": 0,
                "due_soon_count": 0,
                "upcoming_count": 0,
            },
        )
        project["items"].append(item)
        if item["derived_status"] == "Overdue":
            project["overdue_count"] += 1
        if item["derived_status"] == "Due in 7 days":
            project["due_soon_count"] += 1
        if kind == "upcoming" and item["derived_status"] in {"Pending", "Due in 7 days"}:
            project["upcoming_count"] += 1

    sections = list(grouped.values())
    for section in sections:
        section["items"].sort(key=lambda item: item["display_due_date"] or "9999-12-31")
    sections.sort(key=lambda section: section["project_name"].lower())
    return sections


def build_cross_project_calendar(user_id, month_offset=0):
    rows = fetch_user_compliance_rows(user_id)
    approval_lookup = build_project_approval_lookup(fetch_user_approval_rows(user_id))
    event_map = {}
    today = date.today()
    month_start = add_months(today.replace(day=1), month_offset)
    month_end = add_months(month_start, 1) - timedelta(days=1)

    for row in rows:
        item = dict(row)
        project_approvals = approval_lookup.get(item["project_id"], {})
        schedule_dates = generate_schedule_dates(item, project_approvals, month_start, month_end)
        if not schedule_dates:
            due_date = parse_iso_date(item.get("due_date"))
            if due_date and month_start <= due_date <= month_end:
                schedule_dates = [due_date]

        for due_date in schedule_dates:
            due_iso = due_date.isoformat()
            event_kind = "Pending"
            if due_iso < today.isoformat():
                event_kind = "Overdue"
            elif due_iso <= (today + timedelta(days=7)).isoformat():
                event_kind = "Due in 7 days"
            event_map.setdefault(due_iso, []).append(
                {
                    "label": item["project_name"],
                    "meta": item["condition_description"],
                    "kind": event_kind,
                    "href": url_for("approval_detail", project_id=item["project_id"], approval_type=item["approval_type"]),
                }
            )

    month_matrix = calendar.monthcalendar(month_start.year, month_start.month)
    days = []
    for week in month_matrix:
        for day_number in week:
            if day_number == 0:
                days.append({"day": "", "date": "", "events": [], "is_today": False})
                continue
            current_date = date(month_start.year, month_start.month, day_number)
            iso_date = current_date.isoformat()
            days.append(
                {
                    "day": day_number,
                    "date": iso_date,
                    "events": event_map.get(iso_date, []),
                    "is_today": current_date == today,
                }
            )

    return {
        "month_label": f"{calendar.month_name[month_start.month]} {month_start.year}",
        "month_offset": month_offset,
        "days": days,
    }


def build_approval_report_periods(approval):
    issue_date = parse_iso_date(approval["issue_date"])
    expiry_date = parse_iso_date(approval["expiry_date"])
    if not issue_date:
        return []

    periods = []
    if approval["approval_type"] == "EC":
        start_year = issue_date.year
        end_year = (expiry_date or add_months(issue_date, 72)).year
        for year in range(start_year, end_year + 1):
            for month in (6, 12):
                period_date = date(year, month, 1)
                if period_date < issue_date:
                    continue
                if expiry_date and period_date > expiry_date:
                    continue
                periods.append(
                    {
                        "label": period_date.strftime("%b %Y"),
                        "date": period_date.isoformat(),
                        "key": period_date.strftime("%Y%m%d"),
                    }
                )
    else:
        current_date = issue_date
        limit_date = expiry_date or add_months(issue_date, 36)
        while current_date <= limit_date:
            periods.append(
                {
                    "label": current_date.strftime("%b %Y"),
                    "date": current_date.isoformat(),
                    "key": current_date.strftime("%Y%m%d"),
                }
            )
            current_date = add_months(current_date, 6)
    return periods


def annexure_session_code(period_due_date="", period_label=""):
    session_date = parse_iso_date(period_due_date)
    if session_date is None and period_label:
        try:
            session_date = datetime.strptime(period_label, "%b %Y").date()
        except ValueError:
            session_date = None
    if session_date is None:
        return "ANN"

    month_code_map = {
        1: "JA",
        2: "F",
        3: "M",
        4: "A",
        5: "MY",
        6: "J",
        7: "JL",
        8: "AU",
        9: "S",
        10: "O",
        11: "N",
        12: "D",
    }
    return f"{month_code_map.get(session_date.month, 'ANN')}{str(session_date.year)[-2:]}"


def format_annexure_label(annexure_number, period_due_date="", period_label=""):
    if not annexure_number:
        return ""
    return f"{annexure_session_code(period_due_date, period_label)}_Annex {annexure_number}"


def fetch_compliance_report_responses(project_id, approval_type):
    rows = get_db().execute(
        """
        SELECT *
        FROM compliance_report_responses
        WHERE project_id = ? AND approval_type = ?
        ORDER BY period_due_date ASC, compliance_item_id ASC
        """,
        (project_id, approval_type),
    ).fetchall()
    response_map = {}
    for row in rows:
        response = dict(row)
        response["annexure_label"] = format_annexure_label(
            response["annexure_number"], response.get("period_due_date", ""), response["period_label"]
        )
        response_map[(row["compliance_item_id"], row["period_label"])] = response
    return response_map


def next_annexure_number(project_id, approval_type, period_label):
    row = get_db().execute(
        """
        SELECT COALESCE(MAX(annexure_number), 0) AS value
        FROM compliance_report_responses
        WHERE project_id = ? AND approval_type = ? AND period_label = ?
        """,
        (project_id, approval_type, period_label),
    ).fetchone()
    return (row["value"] if row else 0) + 1


def build_approval_documents(project_id, approval_type):
    action_documents = [
        {**dict(document), "source": "Action to be taken"}
        for document in fetch_project_documents(project_id, approval_type=approval_type)
    ]
    report_rows = get_db().execute(
        """
        SELECT crr.*, ci.condition_description
        FROM compliance_report_responses crr
        JOIN compliance_items ci ON ci.id = crr.compliance_item_id
        WHERE crr.project_id = ? AND crr.approval_type = ? AND crr.attachment_stored_filename != ''
        ORDER BY crr.updated_at DESC
        """,
        (project_id, approval_type),
    ).fetchall()
    report_documents = []
    for row in report_rows:
        document = dict(row)
        document["source"] = "Compliance report"
        document["annexure_label"] = format_annexure_label(
            row["annexure_number"], row.get("period_due_date", ""), row["period_label"]
        )
        document["document_title"] = f"{document['annexure_label']} · {row['period_label']}"
        document["original_filename"] = row["attachment_original_filename"]
        report_documents.append(document)
    return action_documents, report_documents


def build_report_export_workbook(project, approval, periods, items, responses, selected_keys):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{approval['approval_type']} report"
    sheet.append([f"{approval['approval_type']} Compliance Report"])
    sheet.append([project["name"], project["location"]])
    sheet.append([])

    selected_periods = [period for period in periods if period["key"] in selected_keys]
    header = ["Condition"] + [period["label"] for period in selected_periods]
    sheet.append(header)
    for item in items:
        row = [item["condition_description"]]
        for period in selected_periods:
            response = responses.get((item["id"], period["label"]), {})
            text_value = response.get("response_text", "")
            if response.get("annexure_label"):
                text_value = f"{text_value}\n{response['annexure_label']}".strip()
            row.append(text_value)
        sheet.append(row)
    return workbook


def build_report_pdf(project, approval, periods, items, responses, selected_keys):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"{approval['approval_type']} Compliance Report", styles["Title"]),
        Paragraph(f"{project['name']} · {project['location']}", styles["Normal"]),
        Spacer(1, 12),
    ]

    selected_periods = [period for period in periods if period["key"] in selected_keys]
    table_data = [["Condition", *[period["label"] for period in selected_periods]]]
    for item in items:
        row = [Paragraph(item["condition_description"], styles["BodyText"])]
        for period in selected_periods:
            response = responses.get((item["id"], period["label"]), {})
            text_value = response.get("response_text", "") or "-"
            if response.get("annexure_label"):
                text_value = f"{text_value}<br/>{response['annexure_label']}"
            row.append(Paragraph(text_value, styles["BodyText"]))
        table_data.append(row)

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9f5ef")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1a6b4a")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d9d9d0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fbfbfa")]),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def record_history(
    project_id,
    item_id,
    change_type,
    field_label="",
    previous_value="",
    new_value="",
    item_snapshot="",
    actor_email="",
):
    get_db().execute(
        """
        INSERT INTO compliance_item_history (
            compliance_item_id,
            project_id,
            change_type,
            field_label,
            previous_value,
            new_value,
            item_snapshot,
            actor_email
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            item_id,
            project_id,
            change_type,
            field_label,
            previous_value,
            new_value,
            item_snapshot,
            actor_email or current_user_email(),
        ),
    )


def sanitize_schedule_source(schedule_source, approval_options):
    if not schedule_source:
        return ""
    return schedule_source if schedule_source in approval_options else ""


def validate_compliance_payload(payload, approval_options):
    if not payload["condition_description"]:
        return "Condition description is required."
    if payload["approval_type"] not in approval_options:
        return "Choose a valid NOC / approval."
    if payload["frequency"] not in FREQUENCY_TYPES:
        return "Invalid frequency."
    if payload["schedule_source"] and payload["schedule_source"] not in approval_options:
        return "Invalid schedule source."
    if payload["status"] not in LIFECYCLE_STATUSES:
        return "Invalid status."
    if is_recurring_frequency(payload["frequency"]) and payload["status"] == "Completed":
        return "Recurring conditions are completed per occurrence, not as one final item."
    if payload["submission_mode"] and payload["submission_mode"] not in SUBMISSION_MODE_OPTIONS:
        return "Invalid submission mode."
    return None


def get_compliance_form_payload(form, approval_options):
    payload = {
        "approval_type": normalize_approval_type(form.get("approval_type", "").strip(), approval_options),
        "condition_description": form.get("condition_description", "").strip(),
        "action_to_be_taken": form.get("action_to_be_taken", "").strip(),
        "due_date": form.get("due_date", "").strip(),
        "frequency": form.get("frequency", "General").strip(),
        "schedule_source": sanitize_schedule_source(form.get("schedule_source", "").strip(), approval_options),
        "submitted_to": form.get("submitted_to", "").strip(),
        "submission_mode": form.get("submission_mode", "").strip(),
        "responsible_person": form.get("responsible_person", "").strip(),
        "acknowledgment_number": form.get("acknowledgment_number", "").strip(),
        "remarks": form.get("remarks", "").strip(),
        "status": form.get("status", "Pending").strip(),
    }
    return payload, validate_compliance_payload(payload, approval_options)


def compare_item_changes(existing_item, payload):
    field_labels = {
        "approval_type": "Approval type",
        "condition_description": "Condition description",
        "action_to_be_taken": "Action to be taken",
        "due_date": "Due date",
        "frequency": "Frequency",
        "schedule_source": "Schedule source",
        "submitted_to": "Submitted to",
        "submission_mode": "Submission mode",
        "responsible_person": "Responsible person",
        "acknowledgment_number": "Acknowledgment number",
        "remarks": "Remarks",
        "status": "Status",
    }
    changes = []
    for field_name, label in field_labels.items():
        previous_value = existing_item[field_name] or ""
        new_value = payload[field_name] or ""
        if previous_value != new_value:
            changes.append((label, previous_value, new_value))
    return changes


def selected_project_approval_types(project_id):
    return [row["approval_type"] for row in fetch_project_approvals(project_id)]


def approval_context_for_template(approval_rows):
    return [
        {
            "approval_type": row["approval_type"],
            "issue_date": row["issue_date"],
            "expiry_date": row["expiry_date"],
        }
        for row in approval_rows
    ]


def merge_ec_approval_notes(existing_notes, metadata):
    parts = []
    if metadata.get("reference_number"):
        parts.append(f"Reference: {metadata['reference_number']}")
    if metadata.get("validity_text"):
        parts.append(f"Validity: {metadata['validity_text']}")
    if metadata.get("proponent_name"):
        parts.append(f"Proponent: {metadata['proponent_name']}")
    if metadata.get("location_text"):
        parts.append(f"Location: {metadata['location_text']}")
    extracted_notes = "\n".join(parts).strip()
    if not extracted_notes:
        return existing_notes or ""
    if not existing_notes:
        return extracted_notes
    if extracted_notes in existing_notes:
        return existing_notes
    return f"{existing_notes}\n{extracted_notes}".strip()


def validate_project_form(name, client_name, location, approval_entries):
    if not name:
        return "Project name is required."
    if not client_name:
        return "Client name is required."
    if not location:
        return "Location is required."
    if not approval_entries:
        return "Select at least one approval type."
    return None


def build_export_workbook(project, approvals, compliance_items):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Compliance Register"

    sheet.append(["Project", project["name"]])
    sheet.append(["Client", project["client_name"]])
    sheet.append(["Location", project["location"]])
    sheet.append([])
    sheet.append(["Selected approvals"])
    sheet.append(["Approval", "Issue date", "Expiry date", "Notes"])
    for approval in approvals:
        sheet.append(
            [
                approval["approval_type"],
                approval["issue_date"],
                approval["expiry_date"],
                approval["approval_notes"],
            ]
        )

    sheet.append([])
    sheet.append(
        [
            "Approval",
            "Condition description",
            "Action to be taken",
            "Status",
            "Urgency",
            "Next due",
            "Base due date",
            "Frequency",
            "Schedule source",
            "Submitted to",
            "Submission mode",
            "Responsible person",
            "Acknowledgment number",
            "Remarks",
            "Document count",
        ]
    )
    for item in compliance_items:
        sheet.append(
            [
                item["approval_type"],
                item["condition_description"],
                item["action_to_be_taken"],
                item["status"],
                item["derived_status"],
                item["next_due_date"],
                item["due_date"],
                item["frequency"],
                item["schedule_source"],
                item["submitted_to"],
                item["submission_mode"],
                item["responsible_person"],
                item["acknowledgment_number"],
                item["remarks"],
                item["document_count"],
            ]
        )

    for column in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(length + 2, 14), 36)

    return workbook


@app.route("/")
def index():
    if g.user:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/signup", methods=["GET", "POST"])
@limiter.limit("5 per minute")
def signup():
    if g.user:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")

        error = None
        if not email:
            error = "Email is required."
        elif not password:
            error = "Password is required."
        elif len(password) < 8:
            error = "Password must be at least 8 characters."
        elif password != confirm_password:
            error = "Passwords do not match."

        db = get_db()
        if error is None and db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone():
            error = "An account with this email already exists."

        if error is None:
            db.execute(
                "INSERT INTO users (email, password_hash) VALUES (?, ?)",
                (email, generate_password_hash(password, method="pbkdf2:sha256")),
            )
            db.commit()
            flash("Account created. Please log in.", "success")
            return redirect(url_for("login"))

        flash(error, "error")

    return render_template("auth.html", mode="signup")


@app.route("/login", methods=["GET", "POST"])
@limiter.limit("8 per minute")
def login():
    if g.user:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = get_db().execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()

        if user is None or not check_password_hash(user["password_hash"], password):
            flash("Invalid email or password.", "error")
        else:
            session.clear()
            session["user_id"] = user["id"]
            return redirect(url_for("dashboard"))

    return render_template("auth.html", mode="login")


@app.post("/logout")
@login_required
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    overdue_items, upcoming_items = build_user_reminders(g.user["id"])
    annotated_items = annotate_user_compliance_items(g.user["id"])
    status_counts = build_status_counts(annotated_items)
    project_count = get_db().execute(
        "SELECT COUNT(*) AS count FROM projects WHERE user_id = ?", (g.user["id"],)
    ).fetchone()["count"]
    recent_projects = build_project_cards(g.user["id"])[:5]
    return render_template(
        "dashboard.html",
        project_count=project_count,
        pending_count=status_counts.get("Pending", 0),
        completed_count=status_counts.get("Completed", 0),
        on_hold_count=status_counts.get("On hold", 0),
        not_applicable_count=status_counts.get("Not applicable", 0),
        recent_projects=recent_projects,
        overdue_items=overdue_items[:8],
        upcoming_items=upcoming_items[:8],
        overdue_count=len(overdue_items),
        upcoming_count=len(upcoming_items),
    )


@app.route("/projects")
@login_required
def projects():
    approval_filter = request.args.get("approval_type", "").strip()
    if approval_filter and approval_filter not in APPROVAL_PORTFOLIO_TYPES:
        approval_filter = ""
    if approval_filter:
        portfolio = build_approval_portfolio(g.user["id"], approval_filter)
        return render_template("approval_portfolio.html", approval_filter=approval_filter, portfolio=portfolio)
    rows = build_project_cards(g.user["id"], approval_filter=approval_filter)
    return render_template("projects.html", projects=rows, approval_filter=approval_filter)


@app.route("/overdue")
@login_required
def overdue_view():
    items = [
        item
        for item in annotate_user_compliance_items(g.user["id"])
        if item["status"] == "Pending" and item["derived_status"] == "Overdue"
    ]
    grouped_projects = group_items_for_overview(items, "overdue")
    return render_template(
        "overdue.html",
        grouped_projects=grouped_projects,
        total_count=len(items),
        project_count=len(grouped_projects),
    )


@app.route("/upcoming")
@login_required
def upcoming_view():
    window_days = request.args.get("days", "30")
    if window_days not in {"7", "15", "30", "60"}:
        window_days = "30"
    days = int(window_days)
    cutoff = (date.today() + timedelta(days=days)).isoformat()
    items = [
        item
        for item in annotate_user_compliance_items(g.user["id"])
        if item["status"] == "Pending"
        and item["display_due_date"]
        and date.today().isoformat() <= item["display_due_date"] <= cutoff
    ]
    grouped_projects = group_items_for_overview(items, "upcoming")
    return render_template(
        "upcoming.html",
        grouped_projects=grouped_projects,
        total_count=len(items),
        project_count=len(grouped_projects),
        selected_window=days,
        window_options=[7, 15, 30, 60],
    )


@app.route("/documents")
@login_required
def documents_view():
    documents = fetch_user_documents(g.user["id"])
    return render_template("documents.html", documents=documents)


@app.route("/calendar")
@login_required
def calendar_view():
    month_offset = request.args.get("offset", "0")
    try:
        month_offset = int(month_offset)
    except ValueError:
        month_offset = 0
    month_offset = max(-6, min(month_offset, 6))
    calendar_data = build_cross_project_calendar(g.user["id"], month_offset=month_offset)
    return render_template("calendar.html", calendar_data=calendar_data)


@app.route("/reports")
@login_required
def reports_view():
    projects = build_project_cards(g.user["id"])
    return render_template("reports.html", projects=projects)


@app.route("/projects/new", methods=["GET", "POST"])
@login_required
def new_project():
    approval_form_entries = build_approval_form_entries()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        client_name = request.form.get("client_name", "").strip()
        location = request.form.get("location", "").strip()
        approval_entries, approval_form_entries = parse_approval_form(request.form)

        error = validate_project_form(name, client_name, location, approval_entries)

        if error is None:
            db = get_db()
            cursor = db.execute(
                """
                INSERT INTO projects (user_id, name, client_name, location)
                VALUES (?, ?, ?, ?)
                """,
                (g.user["id"], name, client_name, location),
            )
            project_id = cursor.lastrowid
            db.executemany(
                """
                INSERT INTO project_approvals (project_id, approval_type, issue_date, expiry_date, approval_notes)
                VALUES (?, ?, ?, ?, ?)
                """,
                [
                    (
                        project_id,
                        approval["approval_type"],
                        approval["issue_date"],
                        approval["expiry_date"],
                        approval["approval_notes"],
                    )
                    for approval in approval_entries
                ],
            )
            db.commit()
            flash("Project created successfully.", "success")
            return redirect(url_for("project_detail", project_id=project_id))

        flash(error, "error")

    return render_template(
        "project_form.html",
        approval_form_entries=approval_form_entries,
        project=None,
        form_action=url_for("new_project"),
        form_title="Create a project",
        form_copy="Capture the core details, approvals, dates, and regulatory notes for this compliance workspace.",
        submit_label="Save Project",
    )


@app.route("/projects/<int:project_id>/edit", methods=["GET", "POST"])
@login_required
def edit_project(project_id):
    project = get_owned_project(project_id)
    approval_form_entries = build_approval_form_entries(fetch_project_approvals(project_id))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        client_name = request.form.get("client_name", "").strip()
        location = request.form.get("location", "").strip()
        approval_entries, approval_form_entries = parse_approval_form(request.form)

        error = validate_project_form(name, client_name, location, approval_entries)
        if error is None:
            db = get_db()
            db.execute(
                """
                UPDATE projects
                SET name = ?, client_name = ?, location = ?
                WHERE id = ? AND user_id = ?
                """,
                (name, client_name, location, project_id, g.user["id"]),
            )
            db.execute("DELETE FROM project_approvals WHERE project_id = ?", (project_id,))
            db.executemany(
                """
                INSERT INTO project_approvals (project_id, approval_type, issue_date, expiry_date, approval_notes)
                VALUES (?, ?, ?, ?, ?)
                """,
                [
                    (
                        project_id,
                        approval["approval_type"],
                        approval["issue_date"],
                        approval["expiry_date"],
                        approval["approval_notes"],
                    )
                    for approval in approval_entries
                ],
            )
            db.commit()
            flash("Project updated successfully.", "success")
            return redirect(url_for("project_detail", project_id=project_id))

        flash(error, "error")

    return render_template(
        "project_form.html",
        approval_form_entries=approval_form_entries,
        project=project,
        form_action=url_for("edit_project", project_id=project_id),
        form_title="Edit project",
        form_copy="Update project details, approval dates, and state-specific notes as the compliance scope evolves.",
        submit_label="Update Project",
    )


@app.post("/projects/<int:project_id>/delete")
@login_required
def delete_project(project_id):
    project = get_owned_project(project_id)
    db = get_db()
    db.execute("DELETE FROM projects WHERE id = ? AND user_id = ?", (project_id, g.user["id"]))
    db.commit()
    flash(f"Project '{project['name']}' deleted.", "success")
    return redirect(url_for("projects"))


@app.route("/projects/<int:project_id>")
@login_required
def project_detail(project_id):
    project = get_owned_project(project_id)
    calendar_view = request.args.get("calendar_view", "1")
    try:
        calendar_view = int(calendar_view)
    except ValueError:
        calendar_view = 1
    if calendar_view not in CALENDAR_VIEW_OPTIONS:
        calendar_view = 1

    approvals = build_project_approval_summaries(project_id)
    calendar_months = build_project_calendar(project_id, months=calendar_view)

    return render_template(
        "project_detail.html",
        project=project,
        approvals=approvals,
        calendar_months=calendar_months,
        calendar_view=calendar_view,
        calendar_view_options=CALENDAR_VIEW_OPTIONS,
    )


@app.route("/projects/<int:project_id>/approvals/<approval_type>")
@login_required
def approval_detail(project_id, approval_type):
    project, approval = get_owned_project_approval(project_id, approval_type)
    approval_rows = fetch_project_approvals(project_id)
    approval_lookup = build_approval_lookup(approval_rows)
    approval_options = [row["approval_type"] for row in approval_rows]
    tab = request.args.get("tab", "conditions")
    status_filter = request.args.get("status_filter", "All")
    if tab not in REPORT_TABS:
        tab = "conditions"
    if status_filter not in STATUS_FILTERS:
        status_filter = "All"

    approval_items = annotate_compliance_items(fetch_project_compliance_items(project_id, approval_type), approval_lookup)
    filtered_items = filter_compliance_items(approval_items, status_filter, "All")
    periods = build_approval_report_periods(approval)
    selected_period_keys = set(request.args.getlist("period"))
    visible_periods = [period for period in periods if period["key"] in selected_period_keys]
    responses = fetch_compliance_report_responses(project_id, approval_type)
    action_documents, report_documents = build_approval_documents(project_id, approval_type)
    condition_count = len(approval_items)
    overdue_count = sum(1 for item in approval_items if item["derived_status"] == "Overdue")
    due_soon_count = sum(1 for item in approval_items if item["derived_status"] == "Due in 7 days")

    return render_template(
        "approval_detail.html",
        project=project,
        approval=annotate_approval_rows([approval])[0],
        approvals=annotate_approval_rows(approval_rows),
        active_approval_type=approval_type,
        active_tab=tab,
        approval_items=filtered_items,
        all_approval_items=approval_items,
        status_filter=status_filter,
        status_filters=STATUS_FILTERS,
        documents=action_documents,
        report_documents=report_documents,
        activity=fetch_project_history(project_id, limit=40, approval_type=approval_type),
        periods=periods,
        visible_periods=visible_periods,
        responses=responses,
        selected_period_keys=selected_period_keys,
        approval_options=approval_options,
        approval_context=approval_context_for_template(approval_rows),
        schedule_preview_text=build_schedule_preview("General", "", "", approval_lookup),
        condition_count=condition_count,
        overdue_count=overdue_count,
        due_soon_count=due_soon_count,
    )


@app.post("/projects/<int:project_id>/approvals/EC/extraction")
@login_required
def upload_ec_letter(project_id):
    project, approval = get_owned_project_approval(project_id, "EC")
    file = request.files.get("ec_letter")

    if file is None or not file.filename:
        flash("Choose an EC PDF to extract conditions from.", "error")
        return redirect(url_for("approval_detail", project_id=project_id, approval_type="EC"))
    if Path(file.filename).suffix.lower() != ".pdf":
        flash("Upload a PDF EC letter for extraction.", "error")
        return redirect(url_for("approval_detail", project_id=project_id, approval_type="EC"))

    original_filename = secure_filename(file.filename)
    unique_name = f"{uuid.uuid4().hex}.pdf"
    save_path = EXTRACTION_UPLOAD_DIR / unique_name
    file.save(save_path)

    try:
        raw_text = normalize_pdf_text(extract_pdf_text(save_path))
    except Exception:
        flash("The EC PDF could not be read. Try a digital PDF or upload conditions via spreadsheet for now.", "error")
        return redirect(url_for("approval_detail", project_id=project_id, approval_type="EC"))

    if len(raw_text.strip()) < 40:
        flash("Not enough readable text was found in this PDF. OCR/scanned PDF support will need a later phase.", "error")
        return redirect(url_for("approval_detail", project_id=project_id, approval_type="EC"))

    metadata = extract_ec_metadata(raw_text)
    conditions = extract_condition_candidates(raw_text)
    if not conditions:
        flash("The EC PDF text was read, but no conditions could be identified for preview.", "error")
        return redirect(url_for("approval_detail", project_id=project_id, approval_type="EC"))

    db = get_db()
    cursor = db.execute(
        """
        INSERT INTO ec_extraction_batches (
            project_id, approval_type, source_filename, stored_filename,
            reference_number, issue_date, validity_text, proponent_name, location_text, raw_text, status
        )
        VALUES (?, 'EC', ?, ?, ?, ?, ?, ?, ?, ?, 'draft')
        """,
        (
            project["id"],
            original_filename,
            unique_name,
            metadata["reference_number"],
            metadata["issue_date"],
            metadata["validity_text"],
            metadata["proponent_name"],
            metadata["location_text"],
            raw_text,
        ),
    )
    batch_id = cursor.lastrowid
    db.executemany(
        """
        INSERT INTO ec_extraction_batch_items (batch_id, item_order, condition_description, action_to_be_taken, is_selected)
        VALUES (?, ?, ?, ?, 1)
        """,
        [
            (batch_id, index + 1, condition, build_default_action_from_condition(condition))
            for index, condition in enumerate(conditions)
        ],
    )
    db.commit()
    flash(f"Draft extraction created with {len(conditions)} candidate conditions. Review before import.", "success")
    return redirect(url_for("review_ec_extraction", batch_id=batch_id))


@app.route("/ec-extractions/<int:batch_id>", methods=["GET", "POST"])
@login_required
def review_ec_extraction(batch_id):
    batch = get_owned_ec_extraction_batch(batch_id)
    if batch["status"] == "discarded":
        flash("This extraction draft was discarded.", "error")
        return redirect(url_for("approval_detail", project_id=batch["project_id"], approval_type="EC"))

    if request.method == "POST":
        action = request.form.get("batch_action", "confirm").strip()
        items = fetch_extraction_batch_items(batch_id)
        db = get_db()
        updated_metadata = {
            "reference_number": request.form.get("reference_number", "").strip(),
            "issue_date": request.form.get("issue_date", "").strip(),
            "validity_text": request.form.get("validity_text", "").strip(),
            "proponent_name": request.form.get("proponent_name", "").strip(),
            "location_text": request.form.get("location_text", "").strip(),
        }
        db.execute(
            """
            UPDATE ec_extraction_batches
            SET reference_number = ?, issue_date = ?, validity_text = ?, proponent_name = ?, location_text = ?
            WHERE id = ?
            """,
            (
                updated_metadata["reference_number"],
                updated_metadata["issue_date"],
                updated_metadata["validity_text"],
                updated_metadata["proponent_name"],
                updated_metadata["location_text"],
                batch_id,
            ),
        )

        selected_rows = []
        for item in items:
            condition_description = request.form.get(f"condition_description_{item['id']}", "").strip()
            action_to_be_taken = request.form.get(f"action_to_be_taken_{item['id']}", "").strip()
            is_selected = 1 if request.form.get(f"is_selected_{item['id']}") else 0
            db.execute(
                """
                UPDATE ec_extraction_batch_items
                SET condition_description = ?, action_to_be_taken = ?, is_selected = ?
                WHERE id = ?
                """,
                (condition_description, action_to_be_taken, is_selected, item["id"]),
            )
            if is_selected and condition_description:
                selected_rows.append((condition_description, action_to_be_taken))

        if action == "discard":
            db.execute("UPDATE ec_extraction_batches SET status = 'discarded' WHERE id = ?", (batch_id,))
            db.commit()
            flash("EC extraction draft discarded.", "success")
            return redirect(url_for("approval_detail", project_id=batch["project_id"], approval_type="EC"))

        if not selected_rows:
            db.commit()
            flash("Select at least one extracted condition to import.", "error")
            return redirect(url_for("review_ec_extraction", batch_id=batch_id))

        approval_row = get_db().execute(
            """
            SELECT *
            FROM project_approvals
            WHERE project_id = ? AND approval_type = 'EC'
            """,
            (batch["project_id"],),
        ).fetchone()
        merged_notes = merge_ec_approval_notes(approval_row["approval_notes"] if approval_row else "", updated_metadata)
        db.execute(
            """
            UPDATE project_approvals
            SET issue_date = CASE WHEN issue_date = '' AND ? != '' THEN ? ELSE issue_date END,
                approval_notes = ?
            WHERE project_id = ? AND approval_type = 'EC'
            """,
            (updated_metadata["issue_date"], updated_metadata["issue_date"], merged_notes, batch["project_id"]),
        )

        inserted_count = 0
        for condition_description, action_to_be_taken in selected_rows:
            cursor = db.execute(
                """
                INSERT INTO compliance_items (
                    project_id, approval_type, source_type, source_batch_id,
                    condition_description, action_to_be_taken, due_date, frequency,
                    schedule_source, submitted_to, submission_mode, responsible_person,
                    acknowledgment_number, remarks, status
                )
                VALUES (?, 'EC', 'ec_extraction', ?, ?, ?, '', 'General', '', '', '', '', '', '', 'Pending')
                """,
                (batch["project_id"], batch_id, condition_description, action_to_be_taken),
            )
            record_history(
                batch["project_id"],
                cursor.lastrowid,
                "ec_extracted",
                "EC extraction",
                "",
                "Pending",
                condition_description,
            )
            inserted_count += 1

        db.execute(
            "UPDATE ec_extraction_batches SET status = 'confirmed', confirmed_at = CURRENT_TIMESTAMP WHERE id = ?",
            (batch_id,),
        )
        db.commit()
        flash(f"Imported {inserted_count} EC conditions from the letter preview.", "success")
        return redirect(url_for("approval_detail", project_id=batch["project_id"], approval_type="EC"))

    return render_template(
        "ec_extraction_preview.html",
        batch=batch,
        items=fetch_extraction_batch_items(batch_id),
    )


@app.get("/projects/<int:project_id>/import-sample")
@login_required
def download_project_import_sample(project_id):
    project = get_owned_project(project_id)
    filename = secure_filename(f"{project['name']}-compliance-sample.csv") or "compliance-sample.csv"
    return Response(
        build_sample_import_csv(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/projects/<int:project_id>/bulk-edit")
@login_required
def bulk_edit_project(project_id):
    project = get_owned_project(project_id)
    approval_rows = fetch_project_approvals(project_id)
    approval_lookup = build_approval_lookup(approval_rows)
    items = annotate_compliance_items(fetch_project_compliance_items(project_id), approval_lookup)
    return render_template(
        "bulk_edit.html",
        project=project,
        items=items,
        approval_options=[approval["approval_type"] for approval in approval_rows],
        approval_context=approval_context_for_template(approval_rows),
        frequency_types=FREQUENCY_TYPES,
    )


@app.post("/projects/<int:project_id>/bulk-edit")
@login_required
def save_bulk_edit_project(project_id):
    get_owned_project(project_id)
    db = get_db()
    approval_options = selected_project_approval_types(project_id)
    item_rows = fetch_project_compliance_items(project_id)
    updated_count = 0

    for item in item_rows:
        item_id = item["id"]
        payload = {
            "approval_type": normalize_approval_type(
                request.form.get(f"approval_type_{item_id}", item["approval_type"]).strip(), approval_options
            ),
            "condition_description": item["condition_description"],
            "action_to_be_taken": request.form.get(f"action_to_be_taken_{item_id}", "").strip(),
            "due_date": request.form.get(f"due_date_{item_id}", "").strip(),
            "frequency": request.form.get(f"frequency_{item_id}", item["frequency"]).strip(),
            "schedule_source": sanitize_schedule_source(
                request.form.get(f"schedule_source_{item_id}", "").strip(), approval_options
            ),
            "submitted_to": item["submitted_to"],
            "submission_mode": item["submission_mode"],
            "responsible_person": item["responsible_person"],
            "acknowledgment_number": item["acknowledgment_number"],
            "remarks": item["remarks"],
            "status": request.form.get(f"status_{item_id}", item["status"]).strip(),
        }
        error = validate_compliance_payload(payload, approval_options)
        if error is not None:
            flash(f"Bulk edit stopped on '{item['condition_description'][:50]}': {error}", "error")
            return redirect(url_for("bulk_edit_project", project_id=project_id))

        changes = compare_item_changes(item, payload)
        if not changes:
            continue

        db.execute(
            """
            UPDATE compliance_items
            SET approval_type = ?, action_to_be_taken = ?, due_date = ?, frequency = ?, schedule_source = ?, status = ?
            WHERE id = ?
            """,
            (
                payload["approval_type"],
                payload["action_to_be_taken"],
                payload["due_date"],
                payload["frequency"],
                payload["schedule_source"],
                payload["status"],
                item_id,
            ),
        )
        for label, previous_value, new_value in changes:
            record_history(
                project_id,
                item_id,
                "bulk_edit",
                label,
                previous_value,
                new_value,
                item["condition_description"],
            )
        updated_count += 1

    db.commit()
    flash(f"Updated {updated_count} compliance items in bulk.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.get("/projects/<int:project_id>/export")
@login_required
def export_project_register(project_id):
    project = get_owned_project(project_id)
    approval_rows = fetch_project_approvals(project_id)
    approval_lookup = build_approval_lookup(approval_rows)
    compliance_items = annotate_compliance_items(fetch_project_compliance_items(project_id), approval_lookup)

    workbook = build_export_workbook(project, approval_rows, compliance_items)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = secure_filename(f"{project['name']}-compliance-register.xlsx") or "compliance-register.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/projects/<int:project_id>/compliance")
@login_required
def add_compliance_item(project_id):
    project = get_owned_project(project_id)
    approval_options = selected_project_approval_types(project_id)
    payload, error = get_compliance_form_payload(request.form, approval_options)

    if error is None:
        db = get_db()
        cursor = db.execute(
            """
            INSERT INTO compliance_items (
                project_id,
                approval_type,
                source_type,
                condition_description,
                action_to_be_taken,
                due_date,
                frequency,
                schedule_source,
                submitted_to,
                submission_mode,
                responsible_person,
                acknowledgment_number,
                remarks,
                status
            )
            VALUES (?, ?, 'manual', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                project["id"],
                payload["approval_type"],
                payload["condition_description"],
                payload["action_to_be_taken"],
                payload["due_date"],
                payload["frequency"],
                payload["schedule_source"],
                payload["submitted_to"],
                payload["submission_mode"],
                payload["responsible_person"],
                payload["acknowledgment_number"],
                payload["remarks"],
                payload["status"],
            ),
        )
        record_history(
            project["id"],
            cursor.lastrowid,
            "created",
            "Compliance item",
            "",
            payload["status"],
            payload["condition_description"],
        )
        db.commit()
        flash("Compliance item added.", "success")
    else:
        flash(error, "error")

    approval_type = payload["approval_type"] if error is None else normalize_approval_type(request.form.get("approval_type", ""))
    return redirect(url_for("approval_detail", project_id=project["id"], approval_type=approval_type or approval_options[0]))


@app.post("/projects/<int:project_id>/import")
@login_required
def import_compliance_items(project_id):
    project = get_owned_project(project_id)
    file = request.files.get("import_file")
    approval_options = selected_project_approval_types(project_id)
    fallback_approval_type = normalize_approval_type(request.form.get("approval_type", ""), approval_options) or (
        approval_options[0] if approval_options else ""
    )

    if file is None or not file.filename:
        flash("Choose a CSV or Excel file to import.", "error")
        return redirect(url_for("approval_detail", project_id=project_id, approval_type=fallback_approval_type))

    if not allowed_import_file(file.filename):
        flash("Unsupported import format. Use CSV or XLSX.", "error")
        return redirect(url_for("approval_detail", project_id=project_id, approval_type=fallback_approval_type))

    try:
        rows = parse_import_rows(file)
    except Exception:
        flash("The file could not be read. Check the format and try again.", "error")
        return redirect(url_for("approval_detail", project_id=project_id, approval_type=fallback_approval_type))
    valid_rows = []
    for row in rows:
        if not row["condition_description"]:
            continue
        row["approval_type"] = normalize_approval_type(
            row.get("approval_type") or request.form.get("approval_type", ""), approval_options
        ) or request.form.get("approval_type", "")
        row["schedule_source"] = sanitize_schedule_source(row.get("schedule_source", ""), approval_options)
        error = validate_compliance_payload(row, approval_options)
        if error is None:
            valid_rows.append(row)

    if not valid_rows:
        flash("No valid condition rows found in the uploaded file.", "error")
        return redirect(url_for("approval_detail", project_id=project_id, approval_type=fallback_approval_type))

    db = get_db()
    for row in valid_rows:
        cursor = db.execute(
            """
            INSERT INTO compliance_items (
                project_id,
                approval_type,
                source_type,
                condition_description,
                action_to_be_taken,
                due_date,
                frequency,
                schedule_source,
                submitted_to,
                submission_mode,
                responsible_person,
                acknowledgment_number,
                remarks,
                status
            )
            VALUES (?, ?, 'import', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                project["id"],
                row["approval_type"],
                row["condition_description"],
                row["action_to_be_taken"],
                row["due_date"],
                row["frequency"],
                row["schedule_source"],
                row["submitted_to"],
                row["submission_mode"],
                row["responsible_person"],
                row["acknowledgment_number"],
                row["remarks"],
                row["status"],
            ),
        )
        record_history(
            project["id"],
            cursor.lastrowid,
            "imported",
            "Import",
            "",
            row["status"],
            row["condition_description"],
        )
    db.commit()
    flash(
        f"Imported {len(valid_rows)} compliance items into {project['name']}. Use bulk edit or 'Edit action / due date' to finish scheduling details.",
        "success",
    )
    return redirect(url_for("approval_detail", project_id=project_id, approval_type=fallback_approval_type))


@app.route("/compliance/<int:item_id>/edit", methods=["GET", "POST"])
@login_required
def edit_compliance_item(item_id):
    item = get_owned_compliance_item(item_id)
    approval_rows = fetch_project_approvals(item["project_id"])
    approval_options = [approval["approval_type"] for approval in approval_rows]

    if request.method == "POST":
        payload, error = get_compliance_form_payload(request.form, approval_options)

        if error is None:
            changes = compare_item_changes(item, payload)
            db = get_db()
            db.execute(
                """
                UPDATE compliance_items
                SET approval_type = ?,
                    condition_description = ?,
                    action_to_be_taken = ?,
                    due_date = ?,
                    frequency = ?,
                    schedule_source = ?,
                    submitted_to = ?,
                    submission_mode = ?,
                    responsible_person = ?,
                    acknowledgment_number = ?,
                    remarks = ?,
                    status = ?
                WHERE id = ?
                """,
                (
                    payload["approval_type"],
                    payload["condition_description"],
                    payload["action_to_be_taken"],
                    payload["due_date"],
                    payload["frequency"],
                    payload["schedule_source"],
                    payload["submitted_to"],
                    payload["submission_mode"],
                    payload["responsible_person"],
                    payload["acknowledgment_number"],
                    payload["remarks"],
                    payload["status"],
                    item_id,
                ),
            )
            for label, previous_value, new_value in changes:
                record_history(
                    item["project_id"],
                    item_id,
                    "edited",
                    label,
                    previous_value,
                    new_value,
                    payload["condition_description"],
                )
            db.commit()
            flash("Compliance item updated.", "success")
            return redirect(url_for("approval_detail", project_id=item["project_id"], approval_type=payload["approval_type"]))

        flash(error, "error")

    current_item = get_owned_compliance_item(item_id)
    project_approvals = annotate_approval_rows(approval_rows)
    return render_template(
        "compliance_form.html",
        item=current_item,
        frequency_types=FREQUENCY_TYPES,
        approval_options=approval_options,
        approval_context=approval_context_for_template(project_approvals),
        schedule_preview_text=build_schedule_preview(
            current_item["frequency"],
            current_item["due_date"],
            current_item["schedule_source"],
            build_approval_lookup(project_approvals),
        ),
    )


@app.post("/compliance/<int:item_id>/delete")
@login_required
def delete_compliance_item(item_id):
    item = get_owned_compliance_item(item_id)
    db = get_db()
    record_history(
        item["project_id"],
        item_id,
        "deleted",
        "Compliance item",
        item["status"],
        "",
        item["condition_description"],
    )
    db.execute("DELETE FROM compliance_items WHERE id = ?", (item_id,))
    db.commit()
    flash("Compliance item deleted.", "success")
    return redirect(url_for("approval_detail", project_id=item["project_id"], approval_type=item["approval_type"]))


@app.post("/compliance/<int:item_id>/status")
@login_required
def update_compliance_status(item_id):
    item = get_owned_compliance_item(item_id)
    new_status = request.form.get("status", "").strip()
    if new_status not in LIFECYCLE_STATUSES:
        flash("Invalid status update.", "error")
        return redirect(url_for("approval_detail", project_id=item["project_id"], approval_type=item["approval_type"]))
    if is_recurring_frequency(item["frequency"]) and new_status == "Completed":
        flash("Recurring conditions should be completed on each occurrence instead of closing the full condition.", "error")
        return redirect(url_for("approval_detail", project_id=item["project_id"], approval_type=item["approval_type"]))

    db = get_db()
    db.execute("UPDATE compliance_items SET status = ? WHERE id = ?", (new_status, item_id))
    record_history(
        item["project_id"],
        item_id,
        "status_changed",
        "Status",
        item["status"],
        new_status,
        item["condition_description"],
    )
    db.commit()
    flash("Compliance status updated.", "success")
    return redirect(url_for("approval_detail", project_id=item["project_id"], approval_type=item["approval_type"]))


@app.post("/compliance/<int:item_id>/upload")
@login_required
def upload_document(item_id):
    item = get_owned_compliance_item(item_id)
    file = request.files.get("document")
    document_title = request.form.get("document_title", "").strip()
    version_notes = request.form.get("version_notes", "").strip()

    if file is None or not file.filename:
        flash("Choose a file to upload.", "error")
        return redirect(url_for("approval_detail", project_id=item["project_id"], approval_type=item["approval_type"]))

    if not allowed_file(file.filename):
        flash("Unsupported file type. Upload a PDF or image.", "error")
        return redirect(url_for("approval_detail", project_id=item["project_id"], approval_type=item["approval_type"]))

    original_filename = secure_filename(file.filename)
    unique_name = f"{uuid.uuid4().hex}{Path(original_filename).suffix.lower()}"
    save_path = UPLOAD_DIR / unique_name
    file.save(save_path)

    if not document_title:
        document_title = Path(original_filename).stem

    db = get_db()
    db.execute(
        """
        INSERT INTO documents (compliance_item_id, original_filename, stored_filename, document_title, version_notes)
        VALUES (?, ?, ?, ?, ?)
        """,
        (item_id, original_filename, unique_name, document_title, version_notes),
    )
    record_history(
        item["project_id"],
        item_id,
        "document_uploaded",
        "Document",
        "",
        document_title,
        item["condition_description"],
    )
    db.commit()
    flash("Document uploaded.", "success")
    return redirect(url_for("approval_detail", project_id=item["project_id"], approval_type=item["approval_type"], tab="documents"))


@app.post("/compliance/<int:item_id>/occurrences/<occurrence_date>/status")
@login_required
def update_occurrence_status(item_id, occurrence_date):
    item = get_owned_compliance_item(item_id)
    new_status = request.form.get("status", "Completed").strip()
    if new_status not in LIFECYCLE_STATUSES:
        flash("Invalid occurrence status.", "error")
        return redirect(url_for("approval_detail", project_id=item["project_id"], approval_type=item["approval_type"]))

    db = get_db()
    db.execute(
        """
        INSERT INTO compliance_occurrences (compliance_item_id, occurrence_date, status, completed_at)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(compliance_item_id, occurrence_date)
        DO UPDATE SET status = excluded.status, completed_at = excluded.completed_at
        """,
        (item_id, occurrence_date, new_status, datetime.utcnow().isoformat() if new_status == "Completed" else ""),
    )
    record_history(
        item["project_id"],
        item_id,
        "occurrence_status_changed",
        "Occurrence status",
        occurrence_date,
        new_status,
        item["condition_description"],
    )
    db.commit()
    flash("Occurrence updated.", "success")
    return redirect(url_for("approval_detail", project_id=item["project_id"], approval_type=item["approval_type"]))


@app.post("/projects/<int:project_id>/approvals/<approval_type>/report")
@login_required
def save_compliance_report(project_id, approval_type):
    project, approval = get_owned_project_approval(project_id, approval_type)
    approval_lookup = build_approval_lookup(fetch_project_approvals(project_id))
    items = annotate_compliance_items(fetch_project_compliance_items(project_id, approval_type), approval_lookup)
    periods = build_approval_report_periods(approval)
    period_lookup = {period["key"]: period for period in periods}
    existing = fetch_compliance_report_responses(project_id, approval_type)
    db = get_db()

    for item in items:
        for period in periods:
            field_prefix = f"{item['id']}_{period['key']}"
            response_text = request.form.get(f"response_{field_prefix}", "").strip()
            file = request.files.get(f"attachment_{field_prefix}")
            current = existing.get((item["id"], period["label"]), {})
            attachment_original = current.get("attachment_original_filename", "")
            attachment_stored = current.get("attachment_stored_filename", "")
            annexure_number = current.get("annexure_number", 0)

            if file and file.filename:
                if not allowed_file(file.filename):
                    flash(f"Unsupported file on {item['condition_description'][:40]} / {period['label']}.", "error")
                    return redirect(url_for("approval_detail", project_id=project_id, approval_type=approval_type, tab="compliance-report"))
                original_filename = secure_filename(file.filename)
                unique_name = f"{uuid.uuid4().hex}{Path(original_filename).suffix.lower()}"
                file.save(UPLOAD_DIR / unique_name)
                attachment_original = original_filename
                attachment_stored = unique_name
                if not annexure_number:
                    annexure_number = next_annexure_number(project_id, approval_type, period["label"])
            if not response_text and not attachment_stored and not current:
                continue

            db.execute(
                """
                INSERT INTO compliance_report_responses (
                    project_id, approval_type, compliance_item_id, period_label, period_due_date,
                    response_text, attachment_original_filename, attachment_stored_filename, annexure_number, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(project_id, approval_type, compliance_item_id, period_label)
                DO UPDATE SET
                    response_text = excluded.response_text,
                    attachment_original_filename = excluded.attachment_original_filename,
                    attachment_stored_filename = excluded.attachment_stored_filename,
                    annexure_number = excluded.annexure_number,
                    updated_at = CURRENT_TIMESTAMP
                """,
                (
                    project_id,
                    approval_type,
                    item["id"],
                    period["label"],
                    period["date"],
                    response_text,
                    attachment_original,
                    attachment_stored,
                    annexure_number,
                ),
            )
    db.commit()
    flash(f"{approval_type} compliance report saved.", "success")
    return redirect(url_for("approval_detail", project_id=project_id, approval_type=approval_type, tab="compliance-report"))


@app.get("/projects/<int:project_id>/approvals/<approval_type>/report/export.xlsx")
@login_required
def export_compliance_report_excel(project_id, approval_type):
    project, approval = get_owned_project_approval(project_id, approval_type)
    periods = build_approval_report_periods(approval)
    selected_keys = set(request.args.getlist("period"))
    items = annotate_compliance_items(fetch_project_compliance_items(project_id, approval_type), build_approval_lookup(fetch_project_approvals(project_id)))
    responses = fetch_compliance_report_responses(project_id, approval_type)
    workbook = build_report_export_workbook(project, approval, periods, items, responses, selected_keys)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = secure_filename(f"{project['name']}-{approval_type}-compliance-report.xlsx")
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/projects/<int:project_id>/approvals/<approval_type>/report/export.pdf")
@login_required
def export_compliance_report_pdf(project_id, approval_type):
    project, approval = get_owned_project_approval(project_id, approval_type)
    periods = build_approval_report_periods(approval)
    selected_keys = set(request.args.getlist("period"))
    items = annotate_compliance_items(fetch_project_compliance_items(project_id, approval_type), build_approval_lookup(fetch_project_approvals(project_id)))
    responses = fetch_compliance_report_responses(project_id, approval_type)
    buffer = build_report_pdf(project, approval, periods, items, responses, selected_keys)
    filename = secure_filename(f"{project['name']}-{approval_type}-compliance-report.pdf")
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")


@app.route("/report-documents/<int:response_id>")
@login_required
def view_report_attachment(response_id):
    response = get_db().execute(
        """
        SELECT crr.*, p.user_id
        FROM compliance_report_responses crr
        JOIN projects p ON p.id = crr.project_id
        WHERE crr.id = ? AND p.user_id = ?
        """,
        (response_id, g.user["id"]),
    ).fetchone()
    if response is None or not response["attachment_stored_filename"]:
        abort(404)
    return send_from_directory(
        app.config["UPLOAD_FOLDER"],
        response["attachment_stored_filename"],
        as_attachment=False,
        download_name=response["attachment_original_filename"],
    )


@app.route("/documents/<int:document_id>")
@login_required
def view_document(document_id):
    document = get_db().execute(
        """
        SELECT d.*, ci.project_id, p.user_id
        FROM documents d
        JOIN compliance_items ci ON ci.id = d.compliance_item_id
        JOIN projects p ON p.id = ci.project_id
        WHERE d.id = ? AND p.user_id = ?
        """,
        (document_id, g.user["id"]),
    ).fetchone()
    if document is None:
        abort(404)

    return send_from_directory(
        app.config["UPLOAD_FOLDER"],
        document["stored_filename"],
        as_attachment=False,
        download_name=document["original_filename"],
    )


@app.errorhandler(413)
def file_too_large(_error):
    flash("File is too large. Please keep uploads under 16 MB.", "error")
    return redirect(request.referrer or url_for("dashboard"))


init_db()


if __name__ == "__main__":
    app.run(debug=True)
